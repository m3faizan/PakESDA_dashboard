"""
Pakistan Intelligence Monitor - Backend API
Real-time intelligence dashboard for Pakistan-related information
"""
import os
import asyncio
import csv
import io
import re
from datetime import datetime, timezone, timedelta
from contextlib import asynccontextmanager
from typing import Optional
import httpx
import feedparser
from openpyxl import load_workbook
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from pymongo import MongoClient
from dotenv import load_dotenv

load_dotenv()

MONGO_URL = os.environ.get("MONGO_URL", "mongodb://localhost:27017")
DB_NAME = os.environ.get("DB_NAME", "pakistan_intel")

client = MongoClient(MONGO_URL)
db = client[DB_NAME]

# Collections
news_collection = db["news"]
economic_collection = db["economic_data"]
security_collection = db["security_alerts"]
weather_collection = db["weather_data"]
airport_history_collection = db["airport_history"]
port_history_collection = db["port_history"]

# Create indexes for efficient querying
airport_history_collection.create_index([("airport_code", 1), ("timestamp", -1)])
port_history_collection.create_index([("port_code", 1), ("timestamp", -1)])

# Flight data cache (refreshes every 6 hours)
flight_cache = {
    "islamabad": {"departures": 0, "arrivals": 0, "updated": None},
    "karachi": {"departures": 0, "arrivals": 0, "updated": None},
    "lahore": {"departures": 0, "arrivals": 0, "updated": None},
    "peshawar": {"departures": 0, "arrivals": 0, "updated": None},
    "multan": {"departures": 0, "arrivals": 0, "updated": None}
}

AIRPORTS = {
    "islamabad": {"code": "ISB", "name": "Islamabad"},
    "karachi": {"code": "KHI", "name": "Karachi"},
    "lahore": {"code": "LHE", "name": "Lahore"},
    "peshawar": {"code": "PEW", "name": "Peshawar"},
    "multan": {"code": "MUX", "name": "Multan"}
}

FLIGHTSTATS_BASE = "https://www.flightstats.com/v2/flight-tracker"

# Port data cache (refreshes every 6 hours)
port_cache = {
    "karachi": {"in_port": 0, "arrivals": 0, "departures": 0, "expected": 0, "updated": None},
    "qasim": {"in_port": 0, "arrivals": 0, "departures": 0, "expected": 0, "updated": None},
    "gwadar": {"in_port": 0, "arrivals": 0, "departures": 0, "expected": 0, "updated": None}
}

PORTS = {
    "karachi": {"id": 236, "name": "Karachi", "code": "PKKHI"},
    "qasim": {"id": 343, "name": "Port Qasim", "code": "PKBQM"},
    "gwadar": {"id": 3609, "name": "Gwadar", "code": "PKGWD"}
}

MYSHIPTRACKING_BASE = "https://www.myshiptracking.com/ports"

# Data cache with timestamps
data_cache = {
    "news": {"data": [], "updated": None},
    "energy": {"data": [], "updated": None},
    "economic": {"data": {}, "updated": None},
    "weather": {"data": [], "updated": None},
    "security": {"data": [], "updated": None},
    "remittances": {"data": {}, "updated": None},
    "gold_reserves": {"data": {}, "updated": None},
    "forex_reserves": {"data": {}, "updated": None},
    "current_account": {"data": {}, "updated": None},
    "imports": {"data": {}, "updated": None},
    "exports": {"data": {}, "updated": None},
    "fdi": {"data": {}, "updated": None},
    "gov_debt": {"data": {}, "updated": None},
    "business_environment": {"data": {}, "updated": None},
    "road_advisory": {"data": [], "updated": None},
    "pkr_usd": {"data": {}, "updated": None},
    "psx_data": {"data": {}, "updated": None},
    "cpi_yoy": {"data": {}, "updated": None},
    "cpi_mom": {"data": {}, "updated": None},
    "cpi_yoy_historical": {"data": {}, "updated": None},
    "cpi_mom_historical": {"data": {}, "updated": None},
    "lsm": {"data": {}, "updated": None},
    "lsm_historical": {"data": {}, "updated": None},
    "auto_vehicles": {"data": {}, "updated": None},
    "fertilizer": {"data": {}, "updated": None},
    "spi_weekly": {"data": {}, "updated": None},
    "spi_monthly": {"data": {}, "updated": None},
    "liquid_forex": {"data": {}, "updated": None}
}

PERSISTED_CACHE_FILE = "/app/backend/persisted_sbp_cache.json"


def load_persisted_cache():
    try:
        if os.path.exists(PERSISTED_CACHE_FILE):
            with open(PERSISTED_CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception as e:
        print(f"Error loading persisted cache: {e}")
    return {}


def save_persisted_cache(cache_obj):
    try:
        with open(PERSISTED_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(cache_obj, f)
    except Exception as e:
        print(f"Error saving persisted cache: {e}")


def persist_cache_entry(cache_key: str, data):
    if data is None:
        return
    cache_obj = load_persisted_cache()
    cache_obj[cache_key] = {
        "updated": datetime.now(timezone.utc).isoformat(),
        "data": data
    }
    save_persisted_cache(cache_obj)


def restore_cache_entry(cache_key: str):
    cache_obj = load_persisted_cache()
    entry = cache_obj.get(cache_key)
    if not entry:
        return None, None
    return entry.get("data"), entry.get("updated")


async def refresh_cache_with_persistence(cache_key: str, ttl_seconds: int, fetcher):
    """Refresh in-memory cache safely and restore last persisted value when upstream is rate-limited."""
    should_refresh = False
    if not data_cache[cache_key]["updated"] or not data_cache[cache_key]["data"]:
        should_refresh = True
    else:
        age = (datetime.now(timezone.utc) - data_cache[cache_key]["updated"]).total_seconds()
        should_refresh = age > ttl_seconds

    if should_refresh:
        fetched = await fetcher()
        if fetched:
            data_cache[cache_key]["data"] = fetched
            data_cache[cache_key]["updated"] = datetime.now(timezone.utc)
            persist_cache_entry(cache_key, fetched)
        elif not data_cache[cache_key]["data"]:
            persisted_data, persisted_updated = restore_cache_entry(cache_key)
            if persisted_data:
                data_cache[cache_key]["data"] = persisted_data
                data_cache[cache_key]["updated"] = datetime.fromisoformat(persisted_updated) if persisted_updated else datetime.now(timezone.utc)
    elif not data_cache[cache_key]["data"]:
        persisted_data, persisted_updated = restore_cache_entry(cache_key)
        if persisted_data:
            data_cache[cache_key]["data"] = persisted_data
            data_cache[cache_key]["updated"] = datetime.fromisoformat(persisted_updated) if persisted_updated else datetime.now(timezone.utc)

# NHMP Road Advisory API
NHMP_ADVISORY_URL = "http://cpo.nhmp.gov.pk:6788/api/TravelAdvisory/FilteredAdvisory"

# SBP API endpoints
SBP_API_KEY = "69C3217DDBE2E78290E66D79E07CCFE19EFB1134"
SBP_REMITTANCES_URL = "https://easydata.sbp.org.pk/api/v1/series/TS_GP_BOP_WR_M.WR0340/data"
SBP_GOLD_RESERVES_URL = "https://easydata.sbp.org.pk/api/v1/series/TS_GP_EXT_PAKRES_M.Z00010/data"
SBP_FOREX_RESERVES_URL = "https://easydata.sbp.org.pk/api/v1/series/TS_GP_EXT_PAKRES_M.Z00020/data"
SBP_BANK_RESERVES_URL = "https://easydata.sbp.org.pk/api/v1/series/TS_GP_EXT_PAKRES_M.Z00050/data"
SBP_SBP_RESERVES_URL = "https://easydata.sbp.org.pk/api/v1/series/TS_GP_EXT_PAKRES_M.Z00020/data"
SBP_CURRENT_ACCOUNT_URL = "https://easydata.sbp.org.pk/api/v1/series/TS_GP_BOP_BPM6SUM_M.P00010/data"
SBP_IMPORTS_URL = "https://easydata.sbp.org.pk/api/v1/series/TS_GP_BOP_XMGS_M.P00320/data"
SBP_EXPORTS_URL = "https://easydata.sbp.org.pk/api/v1/series/TS_GP_BOP_XMGS_M.P00170/data"
SBP_PKR_USD_URL = "https://easydata.sbp.org.pk/api/v1/series/TS_GP_ES_FADERPKR_M.XRDAVG0220/data"
SBP_FDI_URL = "https://easydata.sbp.org.pk/api/v1/series/TS_GP_FI_SUMFIPK_M.FI00030/data"
SBP_GOV_DEBT_TOTAL_URL = "https://easydata.sbp.org.pk/api/v1/series/TS_GP_BAM_CENGOVTD_M.CGD00490/data"
SBP_GOV_DEBT_INTERNAL_URL = "https://easydata.sbp.org.pk/api/v1/series/TS_GP_BAM_CENGOVTD_M.CGD00010/data"
SBP_GOV_DEBT_EXTERNAL_URL = "https://easydata.sbp.org.pk/api/v1/series/TS_GP_BAM_CENGOVTD_M.CGD00460/data"

BUSINESS_ENV_SERIES = {
    "epu_4": "TS_GP_MFS_EPUI_M.EPUI4",
    "epu_2": "TS_GP_MFS_EPUI_M.EPUI2",
    "bci": "TS_GP_RL_BCSIND_M.BCI",
    "cbci": "TS_GP_RL_BCSIND_M.CBCI",
    "ebci": "TS_GP_RL_BCSIND_M.EBCI",
    "sector_manufacturing": "TS_GP_RL_BCSIND_M.BCI_IM",
    "sector_construction": "TS_GP_RL_BCSIND_M.BCI_IC",
    "sector_wholesale_retail": "TS_GP_RL_BCSIND_M.BCI_ST",
    "sector_other_services": "TS_GP_RL_BCSIND_M.BCI_SOS",
    "driver_general_current": "TS_GP_RL_BCSIND_M.BCI_O1",
    "driver_general_expected": "TS_GP_RL_BCSIND_M.BCI_O2",
    "driver_inflation_expected": "TS_GP_RL_BCSIND_M.BCI_O4",
    "driver_employment_current": "TS_GP_RL_BCSIND_M.BCI_O7",
    "driver_employment_expected": "TS_GP_RL_BCSIND_M.BCI_O8",
    "driver_credit_current": "TS_GP_RL_BCSIND_M.BCI_O9",
    "driver_credit_expected": "TS_GP_RL_BCSIND_M.BCI_O10"
}

# Liquid Foreign Exchange Reserves (weekly data scraped from PDF)
SBP_LIQUID_FX_PDF_URL = "https://www.sbp.org.pk/ecodata/forex.pdf"

# CPI API endpoints (2016-present - latest base year)
SBP_CPI_YOY_URL = "https://easydata.sbp.org.pk/api/v1/series/TS_GP_PT_CPI_M.P00011516/data"
SBP_CPI_MOM_URL = "https://easydata.sbp.org.pk/api/v1/series/TS_GP_PT_CPI_M.P00461516/data"

# SPI Google Sheet (user-provided)
SPI_GOOGLE_SHEET_ID = "1q2ixfzKUMx9Y5yX_jPUx5B1wda7bY3hefKAiae5Io1k"
SPI_WEEKLY_SHEET_NAME = "main raw data"
SPI_MONTHLY_SHEET_NAME = "monthly_spi"

# Historical CPI API endpoints by base year period
CPI_HISTORICAL_APIS = [
    {
        "url": "https://easydata.sbp.org.pk/api/v1/series/TS_GP_PT_CPI_M.P00011516/data",  # YoY
        "mom_url": "https://easydata.sbp.org.pk/api/v1/series/TS_GP_PT_CPI_M.P00461516/data",  # MoM
        "base_year": "2015-16",
        "start_date": "2016-07-01",
        "end_date": None,
        "label": "Base: 2015-16"
    },
    {
        "yoy_series": "TS_GP_RLS_CPI0708_M.P00010708",
        "mom_series": "TS_GP_RLS_CPI0708_M.P00020708",
        "base_year": "2007-08",
        "start_date": "2008-07-01",
        "end_date": "2016-06-30",
        "label": "Base: 2007-08"
    },
    {
        "yoy_series": "TS_GP_RLS_CPI0001_M.P00010001",
        "mom_series": "TS_GP_RLS_CPI0001_M.P00020001",
        "base_year": "2000-01",
        "start_date": "2001-07-01",
        "end_date": "2008-06-30",
        "label": "Base: 2000-01"
    },
    {
        "yoy_series": "TS_GP_RLS_CPI9091_M.P00019091",
        "mom_series": "TS_GP_RLS_CPI9091_M.P00029091",
        "base_year": "1990-91",
        "start_date": "1995-01-01",
        "end_date": "2001-06-30",
        "label": "Base: 1990-91"
    },
    {
        "yoy_series": "TS_GP_RLS_CPI8081_M.P00018081",
        "mom_series": "TS_GP_RLS_CPI8081_M.P00028081",
        "base_year": "1980-81",
        "start_date": "1985-03-01",
        "end_date": "1994-12-31",
        "label": "Base: 1980-81"
    },
    {
        "yoy_series": "TS_GP_RLS_CPI7576_M.P00017576",
        "mom_series": "TS_GP_RLS_CPI7576_M.P00027576",
        "base_year": "1975-76",
        "start_date": "1982-07-01",
        "end_date": "1985-02-28",
        "label": "Base: 1975-76"
    },
    {
        "yoy_series": "TS_GP_RLS_CPI6970_M.P00016970",
        "mom_series": "TS_GP_RLS_CPI6970_M.P00026970",
        "base_year": "1969-70",
        "start_date": "1974-01-01",
        "end_date": "1982-06-30",
        "label": "Base: 1969-70"
    },
    {
        "yoy_series": "TS_GP_RLS_CPI5960_M.P00015960",
        "mom_series": "TS_GP_RLS_CPI5960_M.P00025960",
        "base_year": "1959-60",
        "start_date": "1964-07-01",
        "end_date": "1973-12-31",
        "label": "Base: 1959-60"
    }
]

LSM_HISTORICAL_APIS = [
    {
        "series": "TS_GP_RL_LSM1516_M.LSM000160000",
        "base_year": "Base 2015-16",
        "label": "2015-16 Base",
        "start_date": "2016-07-01",
        "end_date": "2025-12-31"
    },
    {
        "series": "TS_GP_RL_LSM_M.LSM000160000",
        "base_year": "Base 2005-06",
        "label": "2005-06 Base",
        "start_date": "2007-07-01",
        "end_date": "2022-06-30"
    },
    {
        "series": "TS_GP_RL_LSM9900_M.LSM000160000",
        "base_year": "Base 1999-2000",
        "label": "1999-2000 Base",
        "start_date": "2000-07-01",
        "end_date": "2007-06-30"
    },
    {
        "series": "TS_GP_RL_LSM8081_M.LSM000160000",
        "base_year": "Base 1980-81",
        "label": "1980-81 Base",
        "start_date": "1985-07-01",
        "end_date": "2000-06-30"
    },
    {
        "series": "TS_GP_RL_LSM7576_M.LSM000160000",
        "base_year": "Base 1975-76",
        "label": "1975-76 Base",
        "start_date": "1981-07-01",
        "end_date": "1985-06-30"
    },
    {
        "series": "TS_GP_RL_LSM6970_M.LSM000160000",
        "base_year": "Base 1969-70",
        "label": "1969-70 Base",
        "start_date": "1977-07-01",
        "end_date": "1981-06-30"
    }
]

AUTO_VEHICLE_SERIES = {
    "production_total": "TS_GP_RLS_PSAUTO_M.TAP_001000",
    "sales_total": "TS_GP_RLS_PSAUTO_M.TAS_001000",
    "production": {
        "cars": "TS_GP_RLS_PSAUTO_M.AP_001001",
        "trucks": "TS_GP_RLS_PSAUTO_M.AP_001003",
        "buses": "TS_GP_RLS_PSAUTO_M.AP_001004",
        "jeeps_pickups": "TS_GP_RLS_PSAUTO_M.AP_001005",
        "tractors": "TS_GP_RLS_PSAUTO_M.AP_001006",
        "two_three_wheelers": "TS_GP_RLS_PSAUTO_M.AP_001002"
    },
    "sales": {
        "cars": "TS_GP_RLS_PSAUTO_M.AS_001001",
        "trucks": "TS_GP_RLS_PSAUTO_M.AS_001003",
        "buses": "TS_GP_RLS_PSAUTO_M.AS_001004",
        "jeeps_pickups": "TS_GP_RLS_PSAUTO_M.AS_001005",
        "tractors": "TS_GP_RLS_PSAUTO_M.AS_001006",
        "two_three_wheelers": "TS_GP_RLS_PSAUTO_M.AS_001002"
    }
}

FERTILIZER_SERIES = {
    "total": "TS_GP_RLS_SALEFERT_M.F_001000",
    "urea": "TS_GP_RLS_SALEFERT_M.U_002000",
    "dap": "TS_GP_RLS_SALEFERT_M.D_003000"
}

# RSS feeds for Pakistan news (expanded per user sources)
PAKISTAN_NEWS_FEEDS = [
    {"name": "Dawn", "url": "https://www.dawn.com/feeds/home", "category": "general"},
    {"name": "Geo News", "url": "https://www.geo.tv/rss/1/1", "category": "general"},
    {"name": "The News", "url": "https://www.thenews.pk/", "category": "general"},
    {"name": "ARY News", "url": "https://arynews.tv/feed/", "category": "general"},
    {"name": "Express Tribune", "url": "https://tribune.com.pk/feed/home", "category": "general"},
    {"name": "Dunya News", "url": "https://dunyanews.tv/index.php/en/RSS", "category": "general"},
    {"name": "Business Recorder", "url": "https://www.brecorder.com/feeds/latest-news", "category": "business"},
    {"name": "Daily Times Business", "url": "https://dailytimes.com.pk/category/business/feed/", "category": "business"},
    {"name": "Profit", "url": "https://profit.pakistantoday.com.pk/feed/", "category": "business"},
    {"name": "Pakistan Today", "url": "https://www.pakistantoday.com.pk/feed/", "category": "general"},
    {"name": "Daily Times", "url": "https://dailytimes.com.pk/feed/", "category": "general"},
    {"name": "Samaa TV", "url": "https://www.samaa.tv/rss", "category": "general"},
    {"name": "APP Pakistan", "url": "https://www.app.com.pk", "category": "general"},
    {"name": "Mettis Global", "url": "https://mettisglobal.news/latest", "category": "business"},
    {"name": "ProPakistani", "url": "https://propakistani.pk/feed/", "category": "tech"},
    {"name": "TechJuice", "url": "https://www.techjuice.pk/feed/", "category": "tech"}
]

# Energy specific RSS feed
ENERGY_NEWS_FEED = {"name": "Energy Update", "url": "https://www.energyupdate.com.pk/feed/", "category": "energy"}

EXCLUDED_NEWS_KEYWORDS = [
    "entertainment", "celebrity", "movie", "movies", "music", "cricket", "football",
    "sports", "fashion", "beauty", "fitness", "gaming", "horoscope",
    "astrology", "comedy", "birthday", "wedding", "netflix", "hbo", "marriage", "score"
]

POLITICS_POLICY_OVERRIDE_KEYWORDS = [
    "politic", "policy", "government", "parliament", "senate", "assembly", "election",
    "prime minister", "president", "cabinet", "minister", "supreme court", "high court",
    "imf", "sbp", "budget", "economy", "economic", "finance", "tax", "regulation",
    "security", "defence", "foreign office", "diplomatic", "protest", "governor"
]


def is_relevant_news_article(title: str, summary: str = ""):
    text = f"{title or ''} {summary or ''}".lower()
    if not text.strip():
        return False

    has_excluded = any(keyword in text for keyword in EXCLUDED_NEWS_KEYWORDS)
    if not has_excluded:
        return True

    # Keep sports-like keywords only when clearly tied to politics/policy/economy/security
    has_override = any(keyword in text for keyword in POLITICS_POLICY_OVERRIDE_KEYWORDS)
    return has_override

def parse_date(date_string):
    """Parse various date formats and return datetime object"""
    if not date_string:
        return None
    
    from email.utils import parsedate_to_datetime
    try:
        # Try RFC 2822 format (common in RSS)
        return parsedate_to_datetime(date_string)
    except:
        pass
    
    # Try common formats
    formats = [
        "%Y-%m-%dT%H:%M:%S%z",
        "%Y-%m-%dT%H:%M:%SZ",
        "%Y-%m-%d %H:%M:%S",
        "%a, %d %b %Y %H:%M:%S %z",
        "%a, %d %b %Y %H:%M:%S %Z",
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(date_string, fmt)
        except:
            continue
    
    return None

def is_within_24_hours(date_string):
    """Check if the article is within last 48 hours (to account for timezone differences)"""
    parsed_date = parse_date(date_string)
    if not parsed_date:
        return True  # Include if we can't parse the date
    
    now = datetime.now(timezone.utc)
    # Make parsed_date timezone aware if it isn't
    if parsed_date.tzinfo is None:
        parsed_date = parsed_date.replace(tzinfo=timezone.utc)
    
    time_diff = now - parsed_date
    return time_diff.total_seconds() <= 172800  # 48 hours in seconds for better coverage

async def fetch_rss_feed(feed_info: dict) -> list:
    """Fetch and parse RSS feed"""
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            response = await client.get(feed_info["url"], headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            })
            if response.status_code == 200:
                feed = feedparser.parse(response.text)
                articles = []
                for entry in feed.entries[:10]:
                    title = entry.get("title", "")
                    summary = entry.get("summary", "")[:250] if entry.get("summary") else ""
                    if not is_relevant_news_article(title, summary):
                        continue

                    articles.append({
                        "title": title,
                        "link": entry.get("link", ""),
                        "published": entry.get("published", ""),
                        "summary": summary,
                        "source": feed_info["name"],
                        "category": feed_info["category"]
                    })
                return articles
    except Exception as e:
        print(f"Error fetching {feed_info['name']}: {e}")
    return []

async def fetch_all_news():
    """Fetch news from all RSS feeds including energy - only last 48 hours"""
    # Include energy feed in main news feeds
    all_feeds = PAKISTAN_NEWS_FEEDS + [ENERGY_NEWS_FEED]
    tasks = [fetch_rss_feed(feed) for feed in all_feeds]
    results = await asyncio.gather(*tasks)
    all_news = []
    for articles in results:
        all_news.extend(articles)
    
    # Filter to only include news from last 48 hours and relevant topics only
    recent_news = [
        article for article in all_news
        if is_within_24_hours(article.get("published", ""))
        and is_relevant_news_article(article.get("title", ""), article.get("summary", ""))
    ]

    # De-duplicate by link/title
    deduped = []
    seen = set()
    for article in recent_news:
        dedupe_key = (article.get("link") or article.get("title") or "").strip().lower()
        if not dedupe_key or dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        deduped.append(article)
    
    # Sort by published date if available
    deduped.sort(key=lambda x: x.get("published", ""), reverse=True)
    return deduped[:150]  # Return up to 150 recent news items

async def fetch_energy_news():
    """Fetch energy news from Energy Update Pakistan"""
    articles = await fetch_rss_feed(ENERGY_NEWS_FEED)
    # Filter to last 24 hours
    recent_articles = [article for article in articles if is_within_24_hours(article.get("published", ""))]
    recent_articles.sort(key=lambda x: x.get("published", ""), reverse=True)
    return recent_articles

async def fetch_remittances_data():
    """Fetch remittances data from State Bank of Pakistan API"""
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            # Fetch all available data from 1972 onwards
            response = await client.get(
                SBP_REMITTANCES_URL,
                params={
                    "api_key": SBP_API_KEY,
                    "start_date": "1972-01-01",
                    "end_date": datetime.now(timezone.utc).strftime("%Y-%m-%d")
                }
            )
            
            if response.status_code == 200:
                data = response.json()
                rows = data.get("rows", [])
                
                if len(rows) >= 13:  # Need at least 13 months for YoY
                    # Data is sorted newest first
                    # Each row: [Dataset Name, Series Key, Series Name, Date, Value, Unit, Status, Comments]
                    history = []
                    for row in rows:
                        history.append({
                            "date": row[3],  # Observation Date
                            "value": float(row[4]),  # Observation Value in Million USD
                            "unit": row[5]  # Unit
                        })
                    
                    # Latest and previous month
                    latest = history[0]
                    previous = history[1]
                    
                    # Find YoY comparison (12 months ago)
                    yoy_value = None
                    latest_date = datetime.strptime(latest["date"], "%Y-%m-%d")
                    for item in history:
                        item_date = datetime.strptime(item["date"], "%Y-%m-%d")
                        months_diff = (latest_date.year - item_date.year) * 12 + (latest_date.month - item_date.month)
                        if months_diff == 12:
                            yoy_value = item["value"]
                            break
                    
                    # Calculate MoM change
                    mom_change = ((latest["value"] - previous["value"]) / previous["value"]) * 100
                    
                    # Calculate YoY change
                    yoy_change = None
                    if yoy_value:
                        yoy_change = ((latest["value"] - yoy_value) / yoy_value) * 100
                    
                    month_name = latest_date.strftime("%B %Y")
                    
                    return {
                        "latest": {
                            "value": round(latest["value"], 2),
                            "month": month_name,
                            "date": latest["date"],
                            "unit": "Million USD"
                        },
                        "previous": {
                            "value": round(previous["value"], 2),
                            "date": previous["date"]
                        },
                        "mom_change": round(mom_change, 2),
                        "yoy_change": round(yoy_change, 2) if yoy_change else None,
                        "history": history,  # Full history for chart
                        "source": "State Bank of Pakistan",
                        "name": "Workers' Remittances",
                        "updated": datetime.now(timezone.utc).isoformat()
                    }
    except Exception as e:
        print(f"Error fetching remittances data: {e}")
    
    # Return None if failed
    return None

async def fetch_sbp_reserves_data(url: str, name: str, start_date: str = "1990-01-01"):
    """Generic function to fetch reserves data from State Bank of Pakistan API"""
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            response = await client.get(
                url,
                params={
                    "api_key": SBP_API_KEY,
                    "start_date": start_date,
                    "end_date": datetime.now(timezone.utc).strftime("%Y-%m-%d")
                }
            )
            
            if response.status_code == 200:
                data = response.json()
                rows = data.get("rows", [])
                
                if len(rows) >= 13:  # Need at least 13 months for YoY
                    history = []
                    for row in rows:
                        history.append({
                            "date": row[3],
                            "value": float(row[4]),
                            "unit": row[5]
                        })
                    
                    latest = history[0]
                    previous = history[1]
                    
                    # Find YoY comparison (12 months ago)
                    yoy_value = None
                    latest_date = datetime.strptime(latest["date"], "%Y-%m-%d")
                    for item in history:
                        item_date = datetime.strptime(item["date"], "%Y-%m-%d")
                        months_diff = (latest_date.year - item_date.year) * 12 + (latest_date.month - item_date.month)
                        if months_diff == 12:
                            yoy_value = item["value"]
                            break
                    
                    # Calculate MoM change
                    mom_change = ((latest["value"] - previous["value"]) / previous["value"]) * 100
                    
                    # Calculate YoY change
                    yoy_change = None
                    if yoy_value:
                        yoy_change = ((latest["value"] - yoy_value) / yoy_value) * 100
                    
                    month_name = latest_date.strftime("%B %Y")
                    
                    return {
                        "latest": {
                            "value": round(latest["value"], 2),
                            "month": month_name,
                            "date": latest["date"],
                            "unit": "Million USD"
                        },
                        "previous": {
                            "value": round(previous["value"], 2),
                            "date": previous["date"]
                        },
                        "mom_change": round(mom_change, 2),
                        "yoy_change": round(yoy_change, 2) if yoy_change else None,
                        "history": history,
                        "source": "State Bank of Pakistan",
                        "name": name,
                        "updated": datetime.now(timezone.utc).isoformat()
                    }
    except Exception as e:
        print(f"Error fetching {name} data: {e}")
    
    return None

async def fetch_gold_reserves_data():
    """Fetch gold reserves data from State Bank of Pakistan"""
    return await fetch_sbp_reserves_data(SBP_GOLD_RESERVES_URL, "Gold Reserves", "1990-01-01")

async def fetch_forex_reserves_data():
    """Fetch total forex reserves (SBP + Banks) from State Bank of Pakistan"""
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            # Fetch SBP reserves and Bank reserves in parallel
            sbp_response, bank_response = await asyncio.gather(
                client.get(
                    SBP_SBP_RESERVES_URL,
                    params={
                        "api_key": SBP_API_KEY,
                        "start_date": "1990-01-01",
                        "end_date": datetime.now(timezone.utc).strftime("%Y-%m-%d")
                    }
                ),
                client.get(
                    SBP_BANK_RESERVES_URL,
                    params={
                        "api_key": SBP_API_KEY,
                        "start_date": "1990-01-01",
                        "end_date": datetime.now(timezone.utc).strftime("%Y-%m-%d")
                    }
                )
            )
            
            sbp_data = []
            bank_data = []
            
            if sbp_response.status_code == 200:
                data = sbp_response.json()
                for row in data.get("rows", []):
                    try:
                        value = float(row[4]) if row[4] else 0
                        sbp_data.append({
                            "date": row[3],
                            "value": value
                        })
                    except (ValueError, IndexError):
                        continue
            
            if bank_response.status_code == 200:
                data = bank_response.json()
                for row in data.get("rows", []):
                    try:
                        value = float(row[4]) if row[4] else 0
                        bank_data.append({
                            "date": row[3],
                            "value": value
                        })
                    except (ValueError, IndexError):
                        continue
            
            # Create a combined history with total = SBP + Bank
            # Index by date for easy lookup
            sbp_by_date = {item["date"]: item["value"] for item in sbp_data}
            bank_by_date = {item["date"]: item["value"] for item in bank_data}
            
            # Get all unique dates
            all_dates = sorted(set(sbp_by_date.keys()) | set(bank_by_date.keys()), reverse=True)
            
            combined_history = []
            sbp_history = []
            bank_history = []
            
            for date in all_dates:
                sbp_val = sbp_by_date.get(date, 0)
                bank_val = bank_by_date.get(date, 0)
                total_val = sbp_val + bank_val
                
                combined_history.append({
                    "date": date,
                    "value": round(total_val, 2),
                    "sbp_reserves": round(sbp_val, 2),
                    "bank_reserves": round(bank_val, 2)
                })
                
                sbp_history.append({
                    "date": date,
                    "value": round(sbp_val, 2)
                })
                
                bank_history.append({
                    "date": date,
                    "value": round(bank_val, 2)
                })
            
            if not combined_history:
                return None
            
            latest = combined_history[0]
            previous = combined_history[1] if len(combined_history) > 1 else None
            
            # Calculate MoM change
            mom_change = None
            if previous:
                mom_change = ((latest["value"] - previous["value"]) / previous["value"]) * 100
            
            # Find YoY comparison
            yoy_value = None
            yoy_change = None
            latest_date = datetime.strptime(latest["date"], "%Y-%m-%d")
            for item in combined_history:
                item_date = datetime.strptime(item["date"], "%Y-%m-%d")
                months_diff = (latest_date.year - item_date.year) * 12 + (latest_date.month - item_date.month)
                if months_diff == 12:
                    yoy_value = item["value"]
                    yoy_change = ((latest["value"] - yoy_value) / yoy_value) * 100
                    break
            
            month_name = latest_date.strftime("%B %Y")
            
            return {
                "latest": {
                    "value": latest["value"],
                    "sbp_reserves": latest["sbp_reserves"],
                    "bank_reserves": latest["bank_reserves"],
                    "month": month_name,
                    "date": latest["date"],
                    "dateFormatted": month_name,
                    "unit": "Million USD"
                },
                "previous": {
                    "value": previous["value"] if previous else None,
                    "sbp_reserves": previous["sbp_reserves"] if previous else None,
                    "bank_reserves": previous["bank_reserves"] if previous else None,
                    "date": previous["date"] if previous else None
                },
                "mom_change": round(mom_change, 2) if mom_change is not None else None,
                "yoy_change": round(yoy_change, 2) if yoy_change is not None else None,
                "yoy_value": round(yoy_value, 2) if yoy_value is not None else None,
                "history": combined_history,
                "sbp_history": sbp_history,
                "bank_history": bank_history,
                "source": "State Bank of Pakistan",
                "name": "Total Forex Reserves",
                "breakdown": {
                    "sbp_reserves": {
                        "name": "SBP Reserves",
                        "latest_value": latest["sbp_reserves"],
                        "unit": "Million USD"
                    },
                    "bank_reserves": {
                        "name": "Bank Reserves",
                        "latest_value": latest["bank_reserves"],
                        "unit": "Million USD"
                    }
                },
                "updated": datetime.now(timezone.utc).isoformat()
            }
    except Exception as e:
        print(f"Error fetching forex reserves: {e}")
        import traceback
        traceback.print_exc()
    
    return None

async def fetch_imports_data():
    """Fetch imports data from State Bank of Pakistan"""
    return await fetch_sbp_reserves_data(SBP_IMPORTS_URL, "Imports", "1990-01-01")

async def fetch_exports_data():
    """Fetch exports data from State Bank of Pakistan"""
    return await fetch_sbp_reserves_data(SBP_EXPORTS_URL, "Exports", "1990-01-01")


async def fetch_fdi_data():
    """Fetch net foreign direct investment in Pakistan"""
    try:
        async with httpx.AsyncClient(timeout=20.0) as client:
            response = await client.get(
                SBP_FDI_URL,
                params={
                    "api_key": SBP_API_KEY,
                    "start_date": "1997-07-01",
                    "end_date": datetime.now(timezone.utc).strftime("%Y-%m-%d")
                }
            )

        if response.status_code != 200:
            return None

        rows = response.json().get("rows", [])
        if len(rows) < 2:
            return None

        history = []
        for row in rows:
            try:
                if not row[3] or row[4] is None:
                    continue
                history.append({
                    "date": row[3],
                    "value": float(row[4]),
                    "unit": row[5] if len(row) > 5 else "Million USD"
                })
            except (ValueError, TypeError, IndexError):
                continue

        if len(history) < 2:
            return None

        latest = history[0]
        previous = history[1]
        latest_date = datetime.strptime(latest["date"], "%Y-%m-%d")
        month_name = latest_date.strftime("%B %Y")

        value_diff = latest["value"] - previous["value"]
        denominator = abs(previous["value"]) if previous["value"] else None
        mom_change = (value_diff / denominator) * 100 if denominator else None

        yoy_value = None
        for item in history:
            item_date = datetime.strptime(item["date"], "%Y-%m-%d")
            months_diff = (latest_date.year - item_date.year) * 12 + (latest_date.month - item_date.month)
            if months_diff == 12:
                yoy_value = item["value"]
                break

        yoy_change = None
        if yoy_value is not None and yoy_value != 0:
            yoy_change = ((latest["value"] - yoy_value) / abs(yoy_value)) * 100

        return {
            "latest": {
                "value": round(latest["value"], 2),
                "month": month_name,
                "date": latest["date"],
                "unit": "Million USD"
            },
            "previous": {
                "value": round(previous["value"], 2),
                "date": previous["date"]
            },
            "mom_change": round(mom_change, 2) if mom_change is not None else None,
            "yoy_change": round(yoy_change, 2) if yoy_change is not None else None,
            "flow_direction": "inflow" if latest["value"] >= 0 else "outflow",
            "history": history,
            "source": "State Bank of Pakistan",
            "name": "Foreign Direct Investment",
            "updated": datetime.now(timezone.utc).isoformat()
        }
    except Exception as e:
        print(f"Error fetching FDI data: {e}")

    return None


async def fetch_gov_debt_data():
    """Fetch Central Government Debt (total + internal + external) from SBP EasyData"""
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            total_response, internal_response, external_response = await asyncio.gather(
                client.get(
                    SBP_GOV_DEBT_TOTAL_URL,
                    params={
                        "api_key": SBP_API_KEY,
                        "start_date": "2010-01-01",
                        "end_date": datetime.now(timezone.utc).strftime("%Y-%m-%d")
                    }
                ),
                client.get(
                    SBP_GOV_DEBT_INTERNAL_URL,
                    params={
                        "api_key": SBP_API_KEY,
                        "start_date": "2010-01-01",
                        "end_date": datetime.now(timezone.utc).strftime("%Y-%m-%d")
                    }
                ),
                client.get(
                    SBP_GOV_DEBT_EXTERNAL_URL,
                    params={
                        "api_key": SBP_API_KEY,
                        "start_date": "2010-01-01",
                        "end_date": datetime.now(timezone.utc).strftime("%Y-%m-%d")
                    }
                )
            )

            if not (total_response.status_code == 200 and internal_response.status_code == 200 and external_response.status_code == 200):
                return None

            def parse_rows(rows):
                parsed = []
                for row in rows:
                    try:
                        if not row[3] or row[4] is None:
                            continue
                        parsed.append({
                            "date": row[3],
                            "value": float(row[4])
                        })
                    except (ValueError, IndexError, TypeError):
                        continue
                return parsed

            total_data = parse_rows(total_response.json().get("rows", []))
            internal_data = parse_rows(internal_response.json().get("rows", []))
            external_data = parse_rows(external_response.json().get("rows", []))

            if not total_data:
                return None

            total_by_date = {item["date"]: item["value"] for item in total_data}
            internal_by_date = {item["date"]: item["value"] for item in internal_data}
            external_by_date = {item["date"]: item["value"] for item in external_data}

            all_dates = sorted(set(total_by_date.keys()) | set(internal_by_date.keys()) | set(external_by_date.keys()), reverse=True)

            history = []
            for date in all_dates:
                total_value = total_by_date.get(date)
                if total_value is None:
                    continue

                internal_value = internal_by_date.get(date, 0)
                external_value = external_by_date.get(date, 0)

                history.append({
                    "date": date,
                    "value": round(total_value, 2),
                    "internal_debt": round(internal_value, 2),
                    "external_debt": round(external_value, 2)
                })

            if len(history) < 2:
                return None

            latest = history[0]
            previous = history[1]

            mom_change = ((latest["value"] - previous["value"]) / previous["value"]) * 100 if previous["value"] else None

            yoy_change = None
            latest_date = datetime.strptime(latest["date"], "%Y-%m-%d")
            for item in history:
                item_date = datetime.strptime(item["date"], "%Y-%m-%d")
                months_diff = (latest_date.year - item_date.year) * 12 + (latest_date.month - item_date.month)
                if months_diff == 12 and item["value"]:
                    yoy_change = ((latest["value"] - item["value"]) / item["value"]) * 100
                    break

            month_name = latest_date.strftime("%B %Y")

            return {
                "latest": {
                    "value": latest["value"],
                    "internal_debt": latest["internal_debt"],
                    "external_debt": latest["external_debt"],
                    "month": month_name,
                    "date": latest["date"],
                    "unit": "Billion PKR"
                },
                "previous": {
                    "value": previous["value"],
                    "date": previous["date"]
                },
                "mom_change": round(mom_change, 2) if mom_change is not None else None,
                "yoy_change": round(yoy_change, 2) if yoy_change is not None else None,
                "history": history,
                "breakdown": {
                    "internal_debt": {
                        "name": "Internal Debt",
                        "latest_value": latest["internal_debt"],
                        "unit": "Billion PKR"
                    },
                    "external_debt": {
                        "name": "External Debt",
                        "latest_value": latest["external_debt"],
                        "unit": "Billion PKR"
                    }
                },
                "source": "State Bank of Pakistan",
                "name": "Central Government Debt",
                "updated": datetime.now(timezone.utc).isoformat()
            }
    except Exception as e:
        print(f"Error fetching government debt data: {e}")

    return None


async def fetch_sbp_series_data(series_code: str, start_date: str = "2018-01-01"):
    """Fetch one SBP series and return ascending monthly history."""
    url = f"https://easydata.sbp.org.pk/api/v1/series/{series_code}/data"
    try:
        async with httpx.AsyncClient(timeout=25.0) as client:
            response = await client.get(
                url,
                params={
                    "api_key": SBP_API_KEY,
                    "start_date": start_date,
                    "end_date": datetime.now(timezone.utc).strftime("%Y-%m-%d")
                }
            )

        if response.status_code != 200:
            return []

        rows = response.json().get("rows", [])
        history = []
        for row in rows:
            try:
                if not row[3] or row[4] is None:
                    continue
                history.append({
                    "date": row[3],
                    "value": float(row[4])
                })
            except (IndexError, ValueError, TypeError):
                continue

        history.sort(key=lambda item: item["date"])
        return history
    except Exception as e:
        print(f"Error fetching SBP series {series_code}: {e}")
        return []


def summarize_monthly_series(history, inverse_good=False):
    if not history:
        return {
            "latest": None,
            "previous": None,
            "mom_change": None,
            "trend_signal": "neutral"
        }

    latest = history[-1]
    previous = history[-2] if len(history) > 1 else None
    mom_change = None
    trend_signal = "neutral"

    if previous and previous["value"]:
        mom_change = ((latest["value"] - previous["value"]) / previous["value"]) * 100
        if inverse_good:
            trend_signal = "positive" if mom_change <= 0 else "negative"
        else:
            trend_signal = "positive" if mom_change >= 0 else "negative"

    latest_date = datetime.strptime(latest["date"], "%Y-%m-%d")

    return {
        "latest": {
            "value": round(latest["value"], 2),
            "date": latest["date"],
            "month": latest_date.strftime("%B %Y")
        },
        "previous": {
            "value": round(previous["value"], 2),
            "date": previous["date"]
        } if previous else None,
        "mom_change": round(mom_change, 2) if mom_change is not None else None,
        "trend_signal": trend_signal
    }


async def fetch_business_environment_data():
    """Fetch EPU + Business Confidence dataset slices for Business Environment panel."""
    try:
        fetch_tasks = [
            fetch_sbp_series_data(series_code, "2017-10-01")
            for series_code in BUSINESS_ENV_SERIES.values()
        ]
        results = await asyncio.gather(*fetch_tasks)
        series_histories = {
            key: results[idx]
            for idx, key in enumerate(BUSINESS_ENV_SERIES.keys())
        }

        epu4_summary = summarize_monthly_series(series_histories["epu_4"], inverse_good=True)
        epu2_summary = summarize_monthly_series(series_histories["epu_2"], inverse_good=True)
        bci_summary = summarize_monthly_series(series_histories["bci"])
        cbci_summary = summarize_monthly_series(series_histories["cbci"])
        ebci_summary = summarize_monthly_series(series_histories["ebci"])

        history_map = {}

        def merge_series(series_key, output_key):
            for item in series_histories.get(series_key, []):
                row = history_map.setdefault(item["date"], {"date": item["date"]})
                row[output_key] = round(item["value"], 2)

        merge_series("bci", "bci")
        merge_series("cbci", "cbci")
        merge_series("ebci", "ebci")
        merge_series("epu_4", "epu4")
        merge_series("epu_2", "epu2")
        merge_series("sector_manufacturing", "manufacturing")
        merge_series("sector_construction", "construction")
        merge_series("sector_wholesale_retail", "wholesale_retail")
        merge_series("sector_other_services", "other_services")

        merged_history = sorted(history_map.values(), key=lambda item: item["date"])

        min_confidence_date = "2017-10-01"
        max_confidence_date = cbci_summary["latest"]["date"] if cbci_summary.get("latest") else (
            bci_summary["latest"]["date"] if bci_summary.get("latest") else None
        )

        confidence_history = [
            {
                "date": item["date"],
                "bci": item.get("bci"),
                "cbci": item.get("cbci"),
                "ebci": item.get("ebci"),
                "manufacturing": item.get("manufacturing"),
                "construction": item.get("construction"),
                "wholesale_retail": item.get("wholesale_retail"),
                "other_services": item.get("other_services")
            }
            for item in merged_history
            if (item.get("bci") is not None or item.get("cbci") is not None or item.get("ebci") is not None)
            and item["date"] >= min_confidence_date
            and (max_confidence_date is None or item["date"] <= max_confidence_date)
        ]

        epu_history = [
            {
                "date": item["date"],
                "epu4": item.get("epu4"),
                "epu2": item.get("epu2")
            }
            for item in merged_history
            if item.get("epu4") is not None and item["date"] >= min_confidence_date
        ]

        sector_mappings = [
            ("Manufacturing", "sector_manufacturing"),
            ("Construction", "sector_construction"),
            ("Wholesale & Retail", "sector_wholesale_retail"),
            ("Other Services", "sector_other_services")
        ]

        sectors_latest = []
        for label, key in sector_mappings:
            summary = summarize_monthly_series(series_histories[key])
            sectors_latest.append({
                "name": label,
                "value": summary["latest"]["value"] if summary["latest"] else None,
                "mom_change": summary["mom_change"],
                "trend_signal": summary["trend_signal"]
            })

        drivers = {
            "general_economic": {
                "current": summarize_monthly_series(series_histories["driver_general_current"]),
                "expected": summarize_monthly_series(series_histories["driver_general_expected"])
            },
            "employment": {
                "current": summarize_monthly_series(series_histories["driver_employment_current"]),
                "expected": summarize_monthly_series(series_histories["driver_employment_expected"])
            },
            "demand_for_credit": {
                "current": summarize_monthly_series(series_histories["driver_credit_current"]),
                "expected": summarize_monthly_series(series_histories["driver_credit_expected"])
            },
            "inflation_expectation": {
                "expected": summarize_monthly_series(series_histories["driver_inflation_expected"], inverse_good=True)
            }
        }

        latest_month = cbci_summary["latest"]["month"] if cbci_summary["latest"] else (
            epu4_summary["latest"]["month"] if epu4_summary["latest"] else "N/A"
        )

        return {
            "epu": {
                "headline": epu4_summary,
                "comparison_2_newspapers": epu2_summary,
                "description": "Policy uncertainty index based on Economy, Policy, and Uncertainty coverage in leading Pakistani newspapers. Higher values imply higher uncertainty.",
                "history": epu_history,
                "date_range": f"{epu_history[0]['date']} to {epu_history[-1]['date']}" if epu_history else None
            },
            "confidence": {
                "headline": {
                    "overall": bci_summary,
                    "current": cbci_summary,
                    "expected": ebci_summary
                },
                "sectors": {
                    "latest": sectors_latest
                },
                "drivers": drivers,
                "history": confidence_history,
                "date_range": f"{confidence_history[0]['date']} to {confidence_history[-1]['date']}" if confidence_history else None
            },
            "latest_month": latest_month,
            "coverage": {
                "epu_series_available": 2,
                "bci_series_available": 107,
                "series_used_in_panel": len(BUSINESS_ENV_SERIES)
            },
            "source": "State Bank of Pakistan",
            "methodology_url": "https://www.sbp.org.pk/research/BCS-m.asp",
            "surveys_info_url": "https://www.sbp.org.pk/research/intro.asp",
            "updated": datetime.now(timezone.utc).isoformat()
        }
    except Exception as e:
        print(f"Error fetching business environment data: {e}")
        return None

async def fetch_pkr_usd_data():
    """Fetch PKR/USD exchange rate data from State Bank of Pakistan"""
    try:
        print(f"Fetching PKR/USD from {SBP_PKR_USD_URL}")
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.get(
                SBP_PKR_USD_URL,
                params={
                    "api_key": SBP_API_KEY,
                    "start_date": "2015-01-01",
                    "end_date": datetime.now(timezone.utc).strftime("%Y-%m-%d")
                }
            )
            
            print(f"PKR/USD API response status: {response.status_code}")
            
            if response.status_code == 200:
                data = response.json()
                rows = data.get("rows", [])
                print(f"PKR/USD rows: {len(rows)}")
                
                if len(rows) >= 2:
                    history = []
                    for row in rows:
                        history.append({
                            "date": row[3],
                            "value": float(row[4]),
                            "unit": row[5]
                        })
                    
                    latest = history[0]
                    previous = history[1]
                    
                    # Find YoY comparison (same day last year or closest)
                    yoy_value = None
                    latest_date = datetime.strptime(latest["date"], "%Y-%m-%d")
                    target_date = latest_date - timedelta(days=365)
                    
                    for item in history:
                        item_date = datetime.strptime(item["date"], "%Y-%m-%d")
                        if abs((item_date - target_date).days) <= 7:
                            yoy_value = item["value"]
                            break
                    
                    # Calculate daily change (percentage)
                    daily_change = ((latest["value"] - previous["value"]) / previous["value"]) * 100
                    
                    # Calculate YoY change
                    yoy_change = None
                    if yoy_value:
                        yoy_change = ((latest["value"] - yoy_value) / yoy_value) * 100
                    
                    date_str = latest_date.strftime("%b %d, %Y")
                    
                    return {
                        "latest": {
                            "value": round(latest["value"], 2),
                            "date": latest["date"],
                            "dateFormatted": date_str,
                            "unit": "PKR"
                        },
                        "previous": {
                            "value": round(previous["value"], 2),
                            "date": previous["date"]
                        },
                        "daily_change": round(daily_change, 4),
                        "yoy_change": round(yoy_change, 2) if yoy_change else None,
                        "history": history,
                        "source": "State Bank of Pakistan",
                        "name": "PKR/USD Exchange Rate",
                        "updated": datetime.now(timezone.utc).isoformat()
                    }
    except Exception as e:
        print(f"Error fetching PKR/USD data: {e}")
    
    return None


async def fetch_liquid_forex_data():
    """Fetch Liquid Foreign Exchange Reserves data from historical JSON and latest SBP PDF weekly table."""
    import json
    import os
    from pypdf import PdfReader

    def fetch_latest_from_sbp_pdf():
        """Read latest weekly FX reserves from SBP PDF (forex.pdf)."""
        try:
            url = "https://www.sbp.org.pk/ecodata/forex.pdf"
            async def _download_bytes():
                async with httpx.AsyncClient(timeout=25.0, follow_redirects=True) as client:
                    resp = await client.get(url)
                    if resp.status_code != 200:
                        return None
                    return resp.content

            pdf_bytes = asyncio.run(_download_bytes())
            if not pdf_bytes:
                return None

            reader = PdfReader(io.BytesIO(pdf_bytes))
            extracted_text = "\n".join([(p.extract_text() or "") for p in reader.pages])
            if not extracted_text:
                return None

            # Weekly rows look like: 6-Mar-26 16,341.1 5,257.2 21,598.3
            pattern = re.compile(r"(\d{1,2}-[A-Za-z]{3}-\d{2})\s+([\d,]+\.\d)\s+([\d,]+\.\d)\s+([\d,]+\.\d)")
            matches = pattern.findall(extracted_text)
            if not matches:
                return None

            weekly_rows = []
            for date_str, sbp_str, bank_str, total_str in matches:
                try:
                    date_obj = datetime.strptime(date_str, "%d-%b-%y")
                    weekly_rows.append({
                        "date": date_obj.strftime("%Y-%m-%d"),
                        "sbp_reserves": float(sbp_str.replace(",", "")),
                        "bank_reserves": float(bank_str.replace(",", "")),
                        "total": float(total_str.replace(",", ""))
                    })
                except Exception:
                    continue

            if not weekly_rows:
                return None

            weekly_rows.sort(key=lambda x: x["date"], reverse=True)
            return weekly_rows[0]
        except Exception as e:
            print(f"Error parsing SBP forex PDF: {e}")
            return None
    
    try:
        # Load historical data from JSON file
        json_path = os.path.join(os.path.dirname(__file__), 'liquid_forex_history.json')
        
        with open(json_path, 'r') as f:
            history = json.load(f)
        
        if not history:
            return None
        
        # Sort by date descending (most recent first)
        history.sort(key=lambda x: x["date"], reverse=True)

        # Always check latest weekly point from SBP PDF and merge/override latest row.
        latest_pdf_point = await asyncio.to_thread(fetch_latest_from_sbp_pdf)
        if latest_pdf_point:
            matched_index = next((idx for idx, item in enumerate(history) if item.get("date") == latest_pdf_point["date"]), None)
            if matched_index is not None:
                history[matched_index]["sbp_reserves"] = latest_pdf_point["sbp_reserves"]
                history[matched_index]["bank_reserves"] = latest_pdf_point["bank_reserves"]
                history[matched_index]["total"] = latest_pdf_point["total"]
            else:
                history.insert(0, latest_pdf_point)

            history.sort(key=lambda x: x["date"], reverse=True)
        
        # Add value field for compatibility
        for item in history:
            item["value"] = item["total"]
        
        latest = history[0]
        previous = history[1] if len(history) > 1 else None
        
        # Calculate week-over-week change
        wow_change = None
        wow_change_pct = None
        if previous:
            wow_change = latest["total"] - previous["total"]
            if previous["total"] > 0:
                wow_change_pct = (wow_change / previous["total"]) * 100
        
        # Calculate % change for each data point (for chart)
        history_reversed = list(reversed(history))
        for i, item in enumerate(history_reversed):
            if i == 0:
                item["pct_change"] = 0
            else:
                prev_item = history_reversed[i-1]
                if prev_item["total"] > 0:
                    item["pct_change"] = round(((item["total"] - prev_item["total"]) / prev_item["total"]) * 100, 2)
                else:
                    item["pct_change"] = 0
        
        # Reverse back to descending order
        history = list(reversed(history_reversed))
        
        from datetime import datetime as dt
        latest_date = dt.strptime(latest["date"], "%Y-%m-%d")
        date_formatted = latest_date.strftime("%b %d, %Y")
        
        return {
            "latest": {
                "value": round(latest["total"], 2),
                "sbp_reserves": round(latest["sbp_reserves"], 2),
                "bank_reserves": round(latest["bank_reserves"], 2),
                "date": latest["date"],
                "dateFormatted": date_formatted,
                "unit": "Million USD"
            },
            "previous": {
                "value": round(previous["total"], 2) if previous else None,
                "sbp_reserves": round(previous["sbp_reserves"], 2) if previous else None,
                "bank_reserves": round(previous["bank_reserves"], 2) if previous else None,
                "date": previous["date"] if previous else None
            },
            "wow_change": round(wow_change, 2) if wow_change is not None else None,
            "wow_change_pct": round(wow_change_pct, 2) if wow_change_pct is not None else None,
            "history": history,
            "total_data_points": len(history),
            "date_range": f"{history[-1]['date']} to {history[0]['date']}",
            "source": "State Bank of Pakistan",
            "name": "Liquid Foreign Exchange Reserves",
            "frequency": "Weekly",
            "breakdown": {
                "sbp_reserves": {
                    "name": "Net Reserves with SBP",
                    "latest_value": round(latest["sbp_reserves"], 2),
                    "unit": "Million USD"
                },
                "bank_reserves": {
                    "name": "Net Reserves with Banks",
                    "latest_value": round(latest["bank_reserves"], 2),
                    "unit": "Million USD"
                }
            },
            "updated": datetime.now(timezone.utc).isoformat()
        }
    except Exception as e:
        print(f"Error fetching liquid forex data: {e}")
        import traceback
        traceback.print_exc()
    
    return None


async def fetch_cpi_data(cpi_type: str = "yoy"):
    """Fetch CPI inflation data from State Bank of Pakistan API
    
    Args:
        cpi_type: 'yoy' for Year-on-Year or 'mom' for Month-on-Month
    """
    try:
        url = SBP_CPI_YOY_URL if cpi_type == "yoy" else SBP_CPI_MOM_URL
        name = "CPI (Year-on-Year)" if cpi_type == "yoy" else "CPI (Month-on-Month)"
        
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.get(
                url,
                params={
                    "api_key": SBP_API_KEY,
                    "start_date": "2016-01-01",
                    "end_date": datetime.now(timezone.utc).strftime("%Y-%m-%d")
                }
            )
            
            if response.status_code == 200:
                data = response.json()
                rows = data.get("rows", [])
                
                if len(rows) >= 2:
                    history = []
                    for row in rows:
                        try:
                            value = float(row[4]) if row[4] else 0
                            history.append({
                                "date": row[3],
                                "value": value,
                                "unit": row[5],
                                "base_year": "2015-16"
                            })
                        except (ValueError, IndexError):
                            continue
                    
                    if not history:
                        return None
                    
                    latest = history[0]
                    previous = history[1] if len(history) > 1 else None
                    
                    # Find YoY comparison (12 months ago)
                    yoy_value = None
                    latest_date = datetime.strptime(latest["date"], "%Y-%m-%d")
                    for item in history:
                        item_date = datetime.strptime(item["date"], "%Y-%m-%d")
                        months_diff = (latest_date.year - item_date.year) * 12 + (latest_date.month - item_date.month)
                        if months_diff == 12:
                            yoy_value = item["value"]
                            break
                    
                    # Calculate change from previous month
                    mom_change = None
                    if previous:
                        mom_change = latest["value"] - previous["value"]
                    
                    month_name = latest_date.strftime("%B %Y")
                    
                    return {
                        "latest": {
                            "value": round(latest["value"], 2),
                            "month": month_name,
                            "date": latest["date"],
                            "unit": "Percent"
                        },
                        "previous": {
                            "value": round(previous["value"], 2) if previous else None,
                            "date": previous["date"] if previous else None
                        },
                        "mom_change": round(mom_change, 2) if mom_change is not None else None,
                        "yoy_comparison": round(yoy_value, 2) if yoy_value is not None else None,
                        "history": history,
                        "source": "State Bank of Pakistan",
                        "name": name,
                        "type": cpi_type,
                        "updated": datetime.now(timezone.utc).isoformat()
                    }
    except Exception as e:
        print(f"Error fetching CPI {cpi_type} data: {e}")
    
    return None


async def fetch_cpi_historical_data(cpi_type: str = "yoy"):
    """Fetch complete CPI history from all base year periods (1964-present)
    
    Args:
        cpi_type: 'yoy' for Year-on-Year or 'mom' for Month-on-Month
    """
    all_history = []
    base_year_markers = []
    
    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            # Fetch all historical periods
            for api_config in CPI_HISTORICAL_APIS:
                try:
                    # Determine the series URL
                    if "url" in api_config:
                        # Latest period (2016-present)
                        url = api_config["url"] if cpi_type == "yoy" else api_config.get("mom_url", api_config["url"])
                    else:
                        # Historical periods
                        series_code = api_config["yoy_series"] if cpi_type == "yoy" else api_config["mom_series"]
                        url = f"https://easydata.sbp.org.pk/api/v1/series/{series_code}/data"
                    
                    params = {
                        "api_key": SBP_API_KEY,
                        "start_date": api_config["start_date"]
                    }
                    if api_config["end_date"]:
                        params["end_date"] = api_config["end_date"]
                    else:
                        params["end_date"] = datetime.now(timezone.utc).strftime("%Y-%m-%d")
                    
                    response = await client.get(url, params=params)
                    
                    if response.status_code == 200:
                        data = response.json()
                        rows = data.get("rows", [])
                        
                        period_history = []
                        for row in rows:
                            try:
                                value = float(row[4]) if row[4] else None
                                if value is not None:
                                    period_history.append({
                                        "date": row[3],
                                        "value": value,
                                        "base_year": api_config["base_year"]
                                    })
                            except (ValueError, IndexError):
                                continue
                        
                        if period_history:
                            all_history.extend(period_history)
                            base_year_markers.append({
                                "date": api_config["start_date"],
                                "label": api_config["label"],
                                "base_year": api_config["base_year"]
                            })
                            print(f"Fetched {len(period_history)} points for {api_config['label']}")
                
                except Exception as e:
                    print(f"Error fetching CPI for {api_config.get('label', 'unknown')}: {e}")
                    continue
            
            # Sort all history by date (oldest to newest)
            all_history.sort(key=lambda x: x["date"])
            
            # Remove duplicates (keep first occurrence for overlapping dates)
            seen_dates = {}
            unique_history = []
            for item in all_history:
                if item["date"] not in seen_dates:
                    seen_dates[item["date"]] = True
                    unique_history.append(item)
            
            # Sort base year markers
            base_year_markers.sort(key=lambda x: x["date"])
            
            if unique_history:
                # Get latest value for summary
                latest = unique_history[-1]
                previous = unique_history[-2] if len(unique_history) > 1 else None
                
                latest_date = datetime.strptime(latest["date"], "%Y-%m-%d")
                month_name = latest_date.strftime("%B %Y")
                
                # Calculate change
                mom_change = None
                if previous:
                    mom_change = latest["value"] - previous["value"]
                
                return {
                    "latest": {
                        "value": round(latest["value"], 2),
                        "month": month_name,
                        "date": latest["date"],
                        "unit": "Percent"
                    },
                    "mom_change": round(mom_change, 2) if mom_change is not None else None,
                    "history": unique_history,
                    "base_year_markers": base_year_markers,
                    "total_data_points": len(unique_history),
                    "date_range": f"{unique_history[0]['date']} to {unique_history[-1]['date']}",
                    "source": "State Bank of Pakistan",
                    "name": f"CPI ({'Year-on-Year' if cpi_type == 'yoy' else 'Month-on-Month'})",
                    "type": cpi_type,
                    "updated": datetime.now(timezone.utc).isoformat()
                }
    
    except Exception as e:
        print(f"Error fetching historical CPI {cpi_type} data: {e}")
        import traceback
        traceback.print_exc()
    
    return None


async def fetch_lsm_historical_data():
    """Fetch complete LSM Quantum Index history across base-year periods (1977-present)."""
    all_history = []
    base_year_markers = []

    try:
        async with httpx.AsyncClient(timeout=20.0) as client:
            for api_config in LSM_HISTORICAL_APIS:
                try:
                    series_code = api_config["series"]
                    url = f"https://easydata.sbp.org.pk/api/v1/series/{series_code}/data"

                    response = await client.get(
                        url,
                        params={
                            "api_key": SBP_API_KEY,
                            "start_date": api_config["start_date"],
                            "end_date": api_config["end_date"]
                        }
                    )

                    if response.status_code != 200:
                        continue

                    payload = response.json()
                    if isinstance(payload, dict) and payload.get("error"):
                        print(f"LSM API rate limit/error for {api_config['label']}: {payload.get('error')}")
                        continue

                    rows = payload.get("rows", [])
                    period_history = []
                    for row in rows:
                        try:
                            if not row[3] or row[4] is None:
                                continue
                            period_history.append({
                                "date": row[3],
                                "value": float(row[4]),
                                "base_year": api_config["base_year"]
                            })
                        except (ValueError, IndexError, TypeError):
                            continue

                    if period_history:
                        all_history.extend(period_history)
                        base_year_markers.append({
                            "date": api_config["start_date"],
                            "label": api_config["label"],
                            "base_year": api_config["base_year"]
                        })

                except Exception as e:
                    print(f"Error fetching LSM period {api_config.get('label', 'unknown')}: {e}")
                    continue

        all_history.sort(key=lambda x: x["date"])

        seen_dates = {}
        unique_history = []
        for item in all_history:
            if item["date"] not in seen_dates:
                seen_dates[item["date"]] = True
                unique_history.append(item)

        base_year_markers.sort(key=lambda x: x["date"])

        if not unique_history:
            return None

        for idx, item in enumerate(unique_history):
            if idx == 0:
                item["pct_change"] = 0
            else:
                prev = unique_history[idx - 1]["value"]
                item["pct_change"] = round(((item["value"] - prev) / prev) * 100, 2) if prev else 0

        latest = unique_history[-1]
        previous = unique_history[-2] if len(unique_history) > 1 else None

        latest_date = datetime.strptime(latest["date"], "%Y-%m-%d")
        month_name = latest_date.strftime("%B %Y")

        mom_change = None
        mom_change_pct = None
        if previous:
            mom_change = latest["value"] - previous["value"]
            mom_change_pct = ((latest["value"] - previous["value"]) / previous["value"]) * 100 if previous["value"] else None

        yoy_change = None
        yoy_value = None
        for item in unique_history:
            item_date = datetime.strptime(item["date"], "%Y-%m-%d")
            months_diff = (latest_date.year - item_date.year) * 12 + (latest_date.month - item_date.month)
            if months_diff == 12:
                yoy_value = item["value"]
                yoy_change = ((latest["value"] - item["value"]) / item["value"]) * 100 if item["value"] else None
                break

        return {
            "latest": {
                "value": round(latest["value"], 2),
                "month": month_name,
                "date": latest["date"],
                "base_year": latest.get("base_year"),
                "unit": "Index"
            },
            "previous": {
                "value": round(previous["value"], 2),
                "date": previous["date"]
            } if previous else None,
            "mom_change": round(mom_change, 2) if mom_change is not None else None,
            "mom_change_pct": round(mom_change_pct, 2) if mom_change_pct is not None else None,
            "yoy_change": round(yoy_change, 2) if yoy_change is not None else None,
            "yoy_value": round(yoy_value, 2) if yoy_value is not None else None,
            "history": unique_history,
            "base_year_markers": base_year_markers,
            "total_data_points": len(unique_history),
            "date_range": f"{unique_history[0]['date']} to {unique_history[-1]['date']}",
            "source": "State Bank of Pakistan / PBS",
            "name": "Quantum Index of Large-scale Manufacturing",
            "updated": datetime.now(timezone.utc).isoformat()
        }

    except Exception as e:
        print(f"Error fetching historical LSM data: {e}")

    return None


def build_lsm_summary_from_historical(historical):
    """Build latest LSM summary payload from historical response object."""
    if not historical:
        return None

    return {
        "latest": historical.get("latest"),
        "previous": historical.get("previous"),
        "mom_change": historical.get("mom_change"),
        "mom_change_pct": historical.get("mom_change_pct"),
        "yoy_change": historical.get("yoy_change"),
        "yoy_value": historical.get("yoy_value"),
        "source": historical.get("source"),
        "name": historical.get("name"),
        "updated": historical.get("updated")
    }


async def fetch_lsm_data():
    """Fetch latest LSM summary based on combined historical base-year series."""
    historical = await fetch_lsm_historical_data()
    return build_lsm_summary_from_historical(historical)


async def fetch_auto_vehicle_data():
    """Fetch production and sales of auto vehicles with category breakdowns."""
    try:
        fetch_map = {
            "production_total": AUTO_VEHICLE_SERIES["production_total"],
            "sales_total": AUTO_VEHICLE_SERIES["sales_total"]
        }
        fetch_map.update({f"production_{k}": v for k, v in AUTO_VEHICLE_SERIES["production"].items()})
        fetch_map.update({f"sales_{k}": v for k, v in AUTO_VEHICLE_SERIES["sales"].items()})

        tasks = [fetch_sbp_series_data(series_code, "2010-01-01") for series_code in fetch_map.values()]
        results = await asyncio.gather(*tasks)
        series_data = {key: results[idx] for idx, key in enumerate(fetch_map.keys())}

        category_labels = {
            "cars": "Cars",
            "trucks": "Trucks",
            "buses": "Buses",
            "jeeps_pickups": "Jeeps & Pickups",
            "tractors": "Tractors",
            "two_three_wheelers": "2 & 3 Wheelers"
        }

        def build_mode(mode: str):
            total_key = f"{mode}_total"
            total_history = series_data.get(total_key, [])
            category_keys = list(AUTO_VEHICLE_SERIES[mode].keys())
            category_histories = {cat: series_data.get(f"{mode}_{cat}", []) for cat in category_keys}

            total_by_date = {item["date"]: item["value"] for item in total_history}
            date_set = set(total_by_date.keys())
            for cat in category_keys:
                date_set.update([item["date"] for item in category_histories[cat]])

            merged = []
            for date in sorted(date_set):
                row = {"date": date}
                category_sum = 0
                for cat in category_keys:
                    val = next((item["value"] for item in category_histories[cat] if item["date"] == date), 0)
                    row[cat] = round(val, 2)
                    category_sum += (val or 0)

                row["total"] = round(total_by_date.get(date, category_sum), 2)
                merged.append(row)

            if not merged:
                return None

            for idx, item in enumerate(merged):
                if idx == 0:
                    item["pct_change"] = 0
                else:
                    prev = merged[idx - 1]["total"]
                    item["pct_change"] = round(((item["total"] - prev) / prev) * 100, 2) if prev else 0

            latest = merged[-1]
            previous = merged[-2] if len(merged) > 1 else None
            latest_date = datetime.strptime(latest["date"], "%Y-%m-%d")
            month_label = latest_date.strftime("%B %Y")
            mom_change_pct = (((latest["total"] - previous["total"]) / previous["total"]) * 100) if previous and previous["total"] else None

            return {
                "latest": {
                    "total": latest["total"],
                    "month": month_label,
                    "date": latest["date"]
                },
                "previous": {
                    "total": previous["total"],
                    "date": previous["date"]
                } if previous else None,
                "mom_change_pct": round(mom_change_pct, 2) if mom_change_pct is not None else None,
                "history": merged,
                "categories": [{"key": cat, "label": category_labels[cat]} for cat in category_keys],
                "date_range": f"{merged[0]['date']} to {merged[-1]['date']}"
            }

        production = build_mode("production")
        sales = build_mode("sales")

        if not production and not sales:
            return None

        latest_date = None
        if production and sales:
            latest_date = max(production["latest"]["date"], sales["latest"]["date"])
        elif production:
            latest_date = production["latest"]["date"]
        elif sales:
            latest_date = sales["latest"]["date"]

        latest_month = datetime.strptime(latest_date, "%Y-%m-%d").strftime("%B %Y") if latest_date else "N/A"

        return {
            "latest_month": latest_month,
            "latest_date": latest_date,
            "production": production,
            "sales": sales,
            "source": "State Bank of Pakistan / PBS",
            "name": "Production and Sale of Auto Vehicles",
            "updated": datetime.now(timezone.utc).isoformat()
        }
    except Exception as e:
        print(f"Error fetching auto vehicle data: {e}")
        return None


async def fetch_fertilizer_data():
    """Fetch fertilizer offtake/sales data with stacked categories and total."""
    try:
        total_history, urea_history, dap_history = await asyncio.gather(
            fetch_sbp_series_data(FERTILIZER_SERIES["total"], "2006-07-01"),
            fetch_sbp_series_data(FERTILIZER_SERIES["urea"], "2006-07-01"),
            fetch_sbp_series_data(FERTILIZER_SERIES["dap"], "2006-07-01")
        )

        total_by_date = {item["date"]: item["value"] for item in total_history}
        urea_by_date = {item["date"]: item["value"] for item in urea_history}
        dap_by_date = {item["date"]: item["value"] for item in dap_history}

        all_dates = sorted(set(total_by_date.keys()) | set(urea_by_date.keys()) | set(dap_by_date.keys()))
        if not all_dates:
            return None

        history = []
        for date in all_dates:
            urea_val = urea_by_date.get(date, 0)
            dap_val = dap_by_date.get(date, 0)
            total_val = total_by_date.get(date, urea_val + dap_val)
            history.append({
                "date": date,
                "urea": round(urea_val, 2),
                "dap": round(dap_val, 2),
                "total": round(total_val, 2)
            })

        for idx, item in enumerate(history):
            if idx == 0:
                item["pct_change"] = 0
            else:
                prev = history[idx - 1]["total"]
                item["pct_change"] = round(((item["total"] - prev) / prev) * 100, 2) if prev else 0

        latest = history[-1]
        previous = history[-2] if len(history) > 1 else None
        latest_date = datetime.strptime(latest["date"], "%Y-%m-%d")

        mom_change_pct = (((latest["total"] - previous["total"]) / previous["total"]) * 100) if previous and previous["total"] else None

        return {
            "latest": {
                "total": latest["total"],
                "urea": latest["urea"],
                "dap": latest["dap"],
                "month": latest_date.strftime("%B %Y"),
                "date": latest["date"],
                "unit": "Thousand Metric Ton"
            },
            "previous": {
                "total": previous["total"],
                "date": previous["date"]
            } if previous else None,
            "mom_change_pct": round(mom_change_pct, 2) if mom_change_pct is not None else None,
            "history": history,
            "categories": [
                {"key": "urea", "label": "Urea"},
                {"key": "dap", "label": "DAP"}
            ],
            "date_range": f"{history[0]['date']} to {history[-1]['date']}",
            "source": "State Bank of Pakistan / PBS",
            "name": "Fertilizer Sales/Offtake",
            "updated": datetime.now(timezone.utc).isoformat()
        }
    except Exception as e:
        print(f"Error fetching fertilizer data: {e}")
        return None


def _safe_float(value):
    if value is None:
        return None

    text = str(value).strip().replace(",", "").replace("%", "")
    if not text:
        return None

    try:
        return float(text)
    except (ValueError, TypeError):
        return None


def _parse_date_by_formats(date_text: str, formats):
    if not date_text:
        return None

    cleaned = str(date_text).strip()
    for fmt in formats:
        try:
            return datetime.strptime(cleaned, fmt)
        except ValueError:
            continue

    return None


async def fetch_google_sheet_workbook():
    """Fetch Google Sheet as XLSX so formula-computed latest values are included."""
    workbook_url = f"https://docs.google.com/spreadsheets/d/{SPI_GOOGLE_SHEET_ID}/export?format=xlsx"

    async with httpx.AsyncClient(timeout=45.0) as client:
        response = await client.get(
            workbook_url,
            follow_redirects=True,
            headers={
                "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36"
            }
        )

    if response.status_code != 200:
        raise ValueError(f"Google Sheet XLSX fetch failed with status {response.status_code}")

    return load_workbook(filename=io.BytesIO(response.content), data_only=True)


def _normalize_sheet_name(name: str):
    return str(name).strip().lower().replace("_", " ")


def _get_workbook_sheet(workbook, expected_name: str):
    target = _normalize_sheet_name(expected_name)
    for sheet_name in workbook.sheetnames:
        if _normalize_sheet_name(sheet_name) == target:
            return workbook[sheet_name]

    return workbook[workbook.sheetnames[0]]


async def fetch_spi_weekly_data():
    """Fetch weekly SPI combined index history from workbook tab Main_Raw_Data."""
    workbook = None
    try:
        workbook = await fetch_google_sheet_workbook()
        sheet = _get_workbook_sheet(workbook, SPI_WEEKLY_SHEET_NAME)

        if sheet.max_row < 2:
            return None

        header_row = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]
        field_map = {}
        for idx, header in enumerate(header_row, start=1):
            if not header:
                continue

            lower_header = str(header).strip().lower()
            if lower_header.startswith("week #"):
                field_map["week"] = idx
            elif lower_header.startswith("week ending"):
                field_map["week_ending"] = idx
            elif lower_header.startswith("q1"):
                field_map["q1"] = idx
            elif lower_header.startswith("q2"):
                field_map["q2"] = idx
            elif lower_header.startswith("q3"):
                field_map["q3"] = idx
            elif lower_header.startswith("q4"):
                field_map["q4"] = idx
            elif lower_header.startswith("q5"):
                field_map["q5"] = idx
            elif lower_header.startswith("combined"):
                field_map["combined"] = idx
            elif lower_header.startswith("items"):
                field_map["items"] = idx
            elif lower_header.startswith("increase"):
                field_map["increase"] = idx
            elif lower_header.startswith("decrease"):
                field_map["decrease"] = idx
            elif lower_header.startswith("stable"):
                field_map["stable"] = idx

        history = []
        for row in range(2, sheet.max_row + 1):
            raw_date = sheet.cell(row=row, column=field_map.get("week_ending", 2)).value
            if isinstance(raw_date, datetime):
                date_obj = raw_date
            else:
                date_obj = _parse_date_by_formats(raw_date, ["%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%d/%m/%Y", "%d-%b-%Y"])

            combined_value = _safe_float(sheet.cell(row=row, column=field_map.get("combined", 8)).value)

            if not date_obj or combined_value is None:
                continue

            week_raw = sheet.cell(row=row, column=field_map.get("week", 1)).value
            week_label = str(int(week_raw)) if isinstance(week_raw, float) else str(week_raw or "").strip()
            if week_label and not week_label.upper().startswith("W"):
                week_label = f"W{week_label}"

            q1 = _safe_float(sheet.cell(row=row, column=field_map.get("q1", 3)).value)
            q2 = _safe_float(sheet.cell(row=row, column=field_map.get("q2", 4)).value)
            q3 = _safe_float(sheet.cell(row=row, column=field_map.get("q3", 5)).value)
            q4 = _safe_float(sheet.cell(row=row, column=field_map.get("q4", 6)).value)
            q5 = _safe_float(sheet.cell(row=row, column=field_map.get("q5", 7)).value)

            history.append({
                "date": date_obj.strftime("%Y-%m-%d"),
                "week": week_label,
                "week_ending_formatted": date_obj.strftime("%b %d, %Y"),
                "value": round(combined_value, 2),
                "q1": round(q1, 2) if q1 is not None else None,
                "q2": round(q2, 2) if q2 is not None else None,
                "q3": round(q3, 2) if q3 is not None else None,
                "q4": round(q4, 2) if q4 is not None else None,
                "q5": round(q5, 2) if q5 is not None else None,
                "items_tracked": int(_safe_float(sheet.cell(row=row, column=field_map.get("items", 9)).value) or 0),
                "increase": int(_safe_float(sheet.cell(row=row, column=field_map.get("increase", 10)).value) or 0),
                "decrease": int(_safe_float(sheet.cell(row=row, column=field_map.get("decrease", 11)).value) or 0),
                "stable": int(_safe_float(sheet.cell(row=row, column=field_map.get("stable", 12)).value) or 0)
            })

        history.sort(key=lambda x: x["date"])
        if not history:
            return None

        for idx, item in enumerate(history):
            if idx == 0:
                item["pct_change"] = 0
            else:
                prev = history[idx - 1]["value"]
                item["pct_change"] = round(((item["value"] - prev) / prev) * 100, 2) if prev else 0

        latest = history[-1]
        previous = history[-2] if len(history) > 1 else None
        wow_change = (latest["value"] - previous["value"]) if previous else None
        wow_change_pct = (((wow_change / previous["value"]) * 100) if previous and previous["value"] else None)

        return {
            "latest": {
                "value": latest["value"],
                "week": latest["week"],
                "date": latest["date"],
                "week_ending_formatted": latest["week_ending_formatted"],
                "q1": latest.get("q1"),
                "q2": latest.get("q2"),
                "q3": latest.get("q3"),
                "q4": latest.get("q4"),
                "q5": latest.get("q5"),
                "items_tracked": latest.get("items_tracked", 0),
                "increase": latest.get("increase", 0),
                "decrease": latest.get("decrease", 0),
                "stable": latest.get("stable", 0),
                "unit": "Index Points"
            },
            "previous": {
                "value": previous["value"] if previous else None,
                "date": previous["date"] if previous else None,
                "week": previous["week"] if previous else None
            },
            "primary_change": round(wow_change, 2) if wow_change is not None else None,
            "primary_change_pct": round(wow_change_pct, 2) if wow_change_pct is not None else None,
            "primary_change_label": "WoW",
            "wow_change": round(wow_change, 2) if wow_change is not None else None,
            "wow_change_pct": round(wow_change_pct, 2) if wow_change_pct is not None else None,
            "history": history,
            "available_series": ["value", "q1", "q2", "q3", "q4", "q5"],
            "series_labels": {
                "value": "Combined",
                "q1": "Q1",
                "q2": "Q2",
                "q3": "Q3",
                "q4": "Q4",
                "q5": "Q5"
            },
            "total_data_points": len(history),
            "date_range": f"{history[0]['date']} to {history[-1]['date']}",
            "source": "SPI Pakistan Dashboard",
            "name": "SPI (Weekly Combined)",
            "frequency": "Weekly",
            "updated": datetime.now(timezone.utc).isoformat()
        }
    except Exception as e:
        print(f"Error fetching weekly SPI data: {e}")
    finally:
        if workbook:
            workbook.close()

    return None


async def fetch_spi_monthly_data():
    """Fetch monthly SPI index (Q1) from workbook tab Monthly_SPI."""
    workbook = None
    try:
        workbook = await fetch_google_sheet_workbook()
        sheet = _get_workbook_sheet(workbook, SPI_MONTHLY_SHEET_NAME)

        if sheet.max_row < 2:
            return None

        header_row = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]
        month_col = 1
        spi_col = 2
        for idx, header in enumerate(header_row, start=1):
            if not header:
                continue

            lower_header = str(header).strip().lower()
            if lower_header == "month":
                month_col = idx
            elif lower_header.startswith("spi"):
                spi_col = idx

        history = []
        for row in range(2, sheet.max_row + 1):
            raw_month = sheet.cell(row=row, column=month_col).value
            spi_value = _safe_float(sheet.cell(row=row, column=spi_col).value)

            if isinstance(raw_month, datetime):
                month_obj = raw_month
            else:
                month_obj = _parse_date_by_formats(raw_month, ["%b-%y", "%b-%Y", "%Y-%m", "%Y-%m-%d"])

            if not month_obj or spi_value is None:
                continue

            history.append({
                "date": month_obj.strftime("%Y-%m-01"),
                "month": month_obj.strftime("%b %Y"),
                "value": round(spi_value, 2)
            })

        history.sort(key=lambda x: x["date"])
        if not history:
            return None

        for idx, item in enumerate(history):
            if idx == 0:
                item["pct_change"] = 0
            else:
                prev = history[idx - 1]["value"]
                item["pct_change"] = round(((item["value"] - prev) / prev) * 100, 2) if prev else 0

        latest = history[-1]
        previous = history[-2] if len(history) > 1 else None
        mom_change = (latest["value"] - previous["value"]) if previous else None
        mom_change_pct = (((mom_change / previous["value"]) * 100) if previous and previous["value"] else None)

        return {
            "latest": {
                "value": latest["value"],
                "month": latest["month"],
                "date": latest["date"],
                "unit": "Index Points"
            },
            "previous": {
                "value": previous["value"] if previous else None,
                "month": previous["month"] if previous else None,
                "date": previous["date"] if previous else None
            },
            "primary_change": round(mom_change, 2) if mom_change is not None else None,
            "primary_change_pct": round(mom_change_pct, 2) if mom_change_pct is not None else None,
            "primary_change_label": "MoM",
            "mom_change": round(mom_change, 2) if mom_change is not None else None,
            "mom_change_pct": round(mom_change_pct, 2) if mom_change_pct is not None else None,
            "history": history,
            "total_data_points": len(history),
            "date_range": f"{history[0]['date']} to {history[-1]['date']}",
            "source": "SPI Pakistan Dashboard",
            "name": "SPI Monthly (Q1)",
            "frequency": "Monthly",
            "updated": datetime.now(timezone.utc).isoformat()
        }
    except Exception as e:
        print(f"Error fetching monthly SPI data: {e}")
    finally:
        if workbook:
            workbook.close()

    return None


async def fetch_road_advisory():
    """Fetch road advisory data from NHMP"""
    try:
        print(f"Fetching road advisory from {NHMP_ADVISORY_URL}")
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.get(
                NHMP_ADVISORY_URL,
                params={
                    "network": "All",
                    "road": "All",
                    "section": "All",
                    "filter": "true"
                }
            )
            
            print(f"NHMP API response status: {response.status_code}")
            
            if response.status_code == 200:
                data = response.json()
                advisories = data.get("data", [])
                print(f"NHMP API returned {len(advisories)} advisories")
                
                # Filter to last 15 days (to capture recent advisories)
                cutoff_date = datetime.now(timezone.utc) - timedelta(days=15)
                
                filtered_advisories = []
                for adv in advisories:
                    record_time = adv.get("recordTime")
                    if record_time:
                        try:
                            adv_date = datetime.fromisoformat(record_time.replace("Z", "+00:00"))
                            if adv_date.tzinfo is None:
                                adv_date = adv_date.replace(tzinfo=timezone.utc)
                            
                            if adv_date >= cutoff_date:
                                filtered_advisories.append({
                                    "id": adv.get("id"),
                                    "road": adv.get("road"),
                                    "route": adv.get("route"),
                                    "type": adv.get("type"),
                                    "subType": adv.get("subType"),
                                    "fromPlace": adv.get("fromPlace"),
                                    "toPlace": adv.get("toPlace"),
                                    "fromKm": adv.get("fromKilometer"),
                                    "toKm": adv.get("toKilometer"),
                                    "recordTime": record_time,
                                    "source": adv.get("source"),
                                    "destination": adv.get("destination"),
                                    "remarks": adv.get("remarks"),
                                    "lat": adv.get("fromLat"),
                                    "lng": adv.get("fromLng")
                                })
                        except:
                            continue
                
                # Sort by record time (newest first)
                filtered_advisories.sort(key=lambda x: x.get("recordTime", ""), reverse=True)
                
                return filtered_advisories
    except Exception as e:
        print(f"Error fetching road advisory: {e}")
    
    return []

async def fetch_current_account_data():
    """Fetch current account balance data from State Bank of Pakistan"""
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            response = await client.get(
                SBP_CURRENT_ACCOUNT_URL,
                params={
                    "api_key": SBP_API_KEY,
                    "start_date": "1990-01-01",
                    "end_date": datetime.now(timezone.utc).strftime("%Y-%m-%d")
                }
            )
            
            if response.status_code == 200:
                data = response.json()
                rows = data.get("rows", [])
                
                if len(rows) >= 13:
                    history = []
                    for row in rows:
                        history.append({
                            "date": row[3],
                            "value": float(row[4]),  # Can be negative
                            "unit": row[5]
                        })
                    
                    latest = history[0]
                    previous = history[1]
                    
                    # Find YoY comparison
                    yoy_value = None
                    latest_date = datetime.strptime(latest["date"], "%Y-%m-%d")
                    for item in history:
                        item_date = datetime.strptime(item["date"], "%Y-%m-%d")
                        months_diff = (latest_date.year - item_date.year) * 12 + (latest_date.month - item_date.month)
                        if months_diff == 12:
                            yoy_value = item["value"]
                            break
                    
                    # Calculate change (absolute difference for current account since it can be negative)
                    mom_change = latest["value"] - previous["value"]
                    yoy_change = (latest["value"] - yoy_value) if yoy_value is not None else None
                    
                    month_name = latest_date.strftime("%B %Y")
                    
                    return {
                        "latest": {
                            "value": round(latest["value"], 2),
                            "month": month_name,
                            "date": latest["date"],
                            "unit": "Million USD"
                        },
                        "previous": {
                            "value": round(previous["value"], 2),
                            "date": previous["date"]
                        },
                        "mom_change": round(mom_change, 2),  # Absolute change
                        "yoy_change": round(yoy_change, 2) if yoy_change is not None else None,
                        "history": history,
                        "source": "State Bank of Pakistan",
                        "name": "Current Account Balance",
                        "updated": datetime.now(timezone.utc).isoformat()
                    }
    except Exception as e:
        print(f"Error fetching current account data: {e}")
    
    return None

async def fetch_economic_data():
    """Fetch economic data - using mock data for now"""
    # In production, this would call real APIs like Alpha Vantage, Open Exchange Rates, etc.
    return {
        "pkr_usd": {
            "rate": 278.50,
            "change": -0.25,
            "change_percent": -0.09,
            "updated": datetime.now(timezone.utc).isoformat()
        },
        "pkr_eur": {
            "rate": 302.15,
            "change": 0.45,
            "change_percent": 0.15,
            "updated": datetime.now(timezone.utc).isoformat()
        },
        "pkr_gbp": {
            "rate": 352.80,
            "change": -0.15,
            "change_percent": -0.04,
            "updated": datetime.now(timezone.utc).isoformat()
        },
        "psx_kse100": {
            "value": 112450.25,
            "change": 1250.50,
            "change_percent": 1.12,
            "volume": "485.2M",
            "updated": datetime.now(timezone.utc).isoformat()
        },
        "inflation": {
            "cpi": 11.8,
            "food": 12.5,
            "energy": 15.2,
            "month": "January 2026",
            "updated": datetime.now(timezone.utc).isoformat()
        },
        "remittances": {
            "monthly": 2.85,
            "yearly": 31.2,
            "change_percent": 8.5,
            "unit": "billion USD",
            "updated": datetime.now(timezone.utc).isoformat()
        },
        "forex_reserves": {
            "value": 13.2,
            "change": 0.3,
            "unit": "billion USD",
            "updated": datetime.now(timezone.utc).isoformat()
        }
    }

async def fetch_weather_data():
    """Fetch weather alerts for major Pakistani cities"""
    # Mock weather data for major cities
    cities = [
        {"name": "Karachi", "lat": 24.8607, "lon": 67.0011, "temp": 28, "condition": "Partly Cloudy", "humidity": 65},
        {"name": "Lahore", "lat": 31.5204, "lon": 74.3587, "temp": 18, "condition": "Foggy", "humidity": 85},
        {"name": "Islamabad", "lat": 33.6844, "lon": 73.0479, "temp": 12, "condition": "Clear", "humidity": 55},
        {"name": "Peshawar", "lat": 34.0151, "lon": 71.5249, "temp": 14, "condition": "Clear", "humidity": 50},
        {"name": "Quetta", "lat": 30.1798, "lon": 66.9750, "temp": 8, "condition": "Cold Wave", "humidity": 40},
        {"name": "Multan", "lat": 30.1575, "lon": 71.5249, "temp": 20, "condition": "Sunny", "humidity": 45},
    ]
    return cities

async def fetch_security_data():
    """Fetch security alerts and political developments"""
    # Mock security/political data
    alerts = [
        {
            "type": "political",
            "severity": "medium",
            "title": "National Assembly Session Scheduled",
            "description": "Parliament to convene for budget discussions",
            "region": "Islamabad",
            "timestamp": datetime.now(timezone.utc).isoformat()
        },
        {
            "type": "security",
            "severity": "low",
            "title": "Security Enhanced in Major Cities",
            "description": "Routine security measures for upcoming events",
            "region": "Nationwide",
            "timestamp": datetime.now(timezone.utc).isoformat()
        },
        {
            "type": "diplomatic",
            "severity": "info",
            "title": "Foreign Minister's Visit to China",
            "description": "Bilateral talks on CPEC projects scheduled",
            "region": "International",
            "timestamp": datetime.now(timezone.utc).isoformat()
        },
        {
            "type": "economic",
            "severity": "medium",
            "title": "IMF Review Mission Arrives",
            "description": "Economic review for next tranche release",
            "region": "Islamabad",
            "timestamp": datetime.now(timezone.utc).isoformat()
        }
    ]
    return alerts

async def fetch_airport_flights(airport_key: str):
    """Fetch airport departure and arrival counts from FlightStats"""
    import re
    airport = AIRPORTS.get(airport_key)
    if not airport:
        return
    
    code = airport["code"]
    timestamp = datetime.now(timezone.utc)
    
    async with httpx.AsyncClient(timeout=15.0) as client:
        # Fetch departures
        try:
            dep_url = f"{FLIGHTSTATS_BASE}/departures/{code}"
            response = await client.get(dep_url, headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            })
            if response.status_code == 200:
                match = re.search(r'(\d+)\s*results', response.text)
                if match:
                    flight_cache[airport_key]["departures"] = int(match.group(1))
        except Exception as e:
            print(f"Error fetching {airport_key} departures: {e}")
        
        # Fetch arrivals
        try:
            arr_url = f"{FLIGHTSTATS_BASE}/arrivals/{code}"
            response = await client.get(arr_url, headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            })
            if response.status_code == 200:
                match = re.search(r'(\d+)\s*results', response.text)
                if match:
                    flight_cache[airport_key]["arrivals"] = int(match.group(1))
        except Exception as e:
            print(f"Error fetching {airport_key} arrivals: {e}")
    
    flight_cache[airport_key]["updated"] = timestamp
    
    # Store historical record in MongoDB
    try:
        historical_record = {
            "airport_key": airport_key,
            "airport_code": code,
            "airport_name": airport["name"],
            "departures": flight_cache[airport_key]["departures"],
            "arrivals": flight_cache[airport_key]["arrivals"],
            "timestamp": timestamp,
            "date": timestamp.strftime("%Y-%m-%d"),
            "hour": timestamp.hour
        }
        airport_history_collection.insert_one(historical_record)
        print(f"Stored airport history for {airport_key}: dep={historical_record['departures']}, arr={historical_record['arrivals']}")
    except Exception as e:
        print(f"Error storing airport history for {airport_key}: {e}")

async def fetch_all_airports():
    """Fetch flight data for all airports"""
    tasks = [fetch_airport_flights(key) for key in AIRPORTS.keys()]
    await asyncio.gather(*tasks)

async def fetch_port_data(port_key: str):
    """Fetch port vessel data from MyShipTracking"""
    import re
    port = PORTS.get(port_key)
    if not port:
        return
    
    port_id = port["id"]
    timestamp = datetime.now(timezone.utc)
    url = f"https://www.myshiptracking.com/ports/port-id-{port_id}"
    
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            response = await client.get(
                f"https://www.myshiptracking.com/ports/port-of-{port['name'].lower().replace(' ', '-')}-in-pk-pakistan-id-{port_id}",
                headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
            )
            if response.status_code == 200:
                text = response.text
                
                # Extract Vessels In Port
                match = re.search(r'Vessels In Port.*?(\d+)', text, re.DOTALL)
                if match:
                    port_cache[port_key]["in_port"] = int(match.group(1))
                
                # Extract Arrivals (24h)
                match = re.search(r'Arrivals \(24h\).*?(\d+)', text, re.DOTALL)
                if match:
                    port_cache[port_key]["arrivals"] = int(match.group(1))
                
                # Extract Departures (24h)
                match = re.search(r'Departures \(24h\).*?(\d+)', text, re.DOTALL)
                if match:
                    port_cache[port_key]["departures"] = int(match.group(1))
                
                # Extract Expected Arrivals
                match = re.search(r'Expected Arrivals.*?(\d+)', text, re.DOTALL)
                if match:
                    port_cache[port_key]["expected"] = int(match.group(1))
                
                port_cache[port_key]["updated"] = timestamp
                
                # Store historical record in MongoDB
                try:
                    historical_record = {
                        "port_key": port_key,
                        "port_code": port["code"],
                        "port_name": port["name"],
                        "in_port": port_cache[port_key]["in_port"],
                        "arrivals": port_cache[port_key]["arrivals"],
                        "departures": port_cache[port_key]["departures"],
                        "expected": port_cache[port_key]["expected"],
                        "timestamp": timestamp,
                        "date": timestamp.strftime("%Y-%m-%d"),
                        "hour": timestamp.hour
                    }
                    port_history_collection.insert_one(historical_record)
                    print(f"Stored port history for {port_key}: in_port={historical_record['in_port']}, arr={historical_record['arrivals']}, dep={historical_record['departures']}")
                except Exception as e:
                    print(f"Error storing port history for {port_key}: {e}")
    except Exception as e:
        print(f"Error fetching {port_key} port data: {e}")

async def fetch_all_ports():
    """Fetch data for all ports"""
    tasks = [fetch_port_data(key) for key in PORTS.keys()]
    await asyncio.gather(*tasks)

@asynccontextmanager
async def lifespan(app: FastAPI):
    # Startup: Initialize data
    print("Pakistan Intelligence Monitor starting...")
    try:
        data_cache["news"]["data"] = await fetch_all_news()
        data_cache["news"]["updated"] = datetime.now(timezone.utc)
        data_cache["energy"]["data"] = await fetch_energy_news()
        data_cache["energy"]["updated"] = datetime.now(timezone.utc)
        data_cache["economic"]["data"] = await fetch_economic_data()
        data_cache["economic"]["updated"] = datetime.now(timezone.utc)
        data_cache["weather"]["data"] = await fetch_weather_data()
        data_cache["weather"]["updated"] = datetime.now(timezone.utc)
        data_cache["security"]["data"] = await fetch_security_data()
        data_cache["security"]["updated"] = datetime.now(timezone.utc)
    except Exception as e:
        print(f"Error during startup: {e}")
    yield
    # Shutdown
    print("Pakistan Intelligence Monitor shutting down...")

app = FastAPI(
    title="Pakistan Intelligence Monitor API",
    description="Real-time intelligence dashboard for Pakistan",
    version="1.0.0",
    lifespan=lifespan
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Pydantic models
class HealthResponse(BaseModel):
    status: str
    timestamp: str
    version: str

class NewsItem(BaseModel):
    title: str
    link: str
    published: str
    summary: str
    source: str
    category: str

class EconomicData(BaseModel):
    pkr_usd: dict
    pkr_eur: dict
    pkr_gbp: dict
    psx_kse100: dict
    inflation: dict
    remittances: dict
    forex_reserves: dict

class WeatherData(BaseModel):
    name: str
    lat: float
    lon: float
    temp: int
    condition: str
    humidity: int

class SecurityAlert(BaseModel):
    type: str
    severity: str
    title: str
    description: str
    region: str
    timestamp: str

# API Routes
@app.get("/api/health", response_model=HealthResponse)
async def health_check():
    return {
        "status": "healthy",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "version": "1.0.0"
    }

@app.get("/api/news")
async def get_news():
    """Get latest Pakistan news (last 48 hours)"""
    # Refresh news if older than 5 minutes
    if data_cache["news"]["updated"]:
        age = (datetime.now(timezone.utc) - data_cache["news"]["updated"]).total_seconds()
        if age > 300:
            data_cache["news"]["data"] = await fetch_all_news()
            data_cache["news"]["updated"] = datetime.now(timezone.utc)
    else:
        data_cache["news"]["data"] = await fetch_all_news()
        data_cache["news"]["updated"] = datetime.now(timezone.utc)
    
    return {
        "news": data_cache["news"]["data"],
        "updated": data_cache["news"]["updated"].isoformat() if data_cache["news"]["updated"] else None,
        "count": len(data_cache["news"]["data"]),
        "filter": "last_48_hours"
    }

@app.get("/api/energy")
async def get_energy_news():
    """Get energy sector news from Energy Update Pakistan (last 24 hours)"""
    # Refresh if older than 5 minutes
    if data_cache["energy"]["updated"]:
        age = (datetime.now(timezone.utc) - data_cache["energy"]["updated"]).total_seconds()
        if age > 300:
            data_cache["energy"]["data"] = await fetch_energy_news()
            data_cache["energy"]["updated"] = datetime.now(timezone.utc)
    else:
        data_cache["energy"]["data"] = await fetch_energy_news()
        data_cache["energy"]["updated"] = datetime.now(timezone.utc)
    
    return {
        "news": data_cache["energy"]["data"],
        "updated": data_cache["energy"]["updated"].isoformat() if data_cache["energy"]["updated"] else None,
        "count": len(data_cache["energy"]["data"]),
        "source": "Energy Update Pakistan",
        "filter": "last_24_hours"
    }

@app.get("/api/economic")
async def get_economic_data():
    """Get economic indicators"""
    # Refresh if older than 1 minute
    if data_cache["economic"]["updated"]:
        age = (datetime.now(timezone.utc) - data_cache["economic"]["updated"]).total_seconds()
        if age > 60:
            data_cache["economic"]["data"] = await fetch_economic_data()
            data_cache["economic"]["updated"] = datetime.now(timezone.utc)
    else:
        data_cache["economic"]["data"] = await fetch_economic_data()
        data_cache["economic"]["updated"] = datetime.now(timezone.utc)
    
    return {
        "data": data_cache["economic"]["data"],
        "updated": data_cache["economic"]["updated"].isoformat() if data_cache["economic"]["updated"] else None
    }

@app.get("/api/weather")
async def get_weather():
    """Get weather data for major Pakistani cities"""
    if data_cache["weather"]["updated"]:
        age = (datetime.now(timezone.utc) - data_cache["weather"]["updated"]).total_seconds()
        if age > 600:
            data_cache["weather"]["data"] = await fetch_weather_data()
            data_cache["weather"]["updated"] = datetime.now(timezone.utc)
    else:
        data_cache["weather"]["data"] = await fetch_weather_data()
        data_cache["weather"]["updated"] = datetime.now(timezone.utc)
    
    return {
        "cities": data_cache["weather"]["data"],
        "updated": data_cache["weather"]["updated"].isoformat() if data_cache["weather"]["updated"] else None
    }

@app.get("/api/security")
async def get_security_alerts():
    """Get security and political alerts"""
    if data_cache["security"]["updated"]:
        age = (datetime.now(timezone.utc) - data_cache["security"]["updated"]).total_seconds()
        if age > 300:
            data_cache["security"]["data"] = await fetch_security_data()
            data_cache["security"]["updated"] = datetime.now(timezone.utc)
    else:
        data_cache["security"]["data"] = await fetch_security_data()
        data_cache["security"]["updated"] = datetime.now(timezone.utc)
    
    return {
        "alerts": data_cache["security"]["data"],
        "updated": data_cache["security"]["updated"].isoformat() if data_cache["security"]["updated"] else None,
        "count": len(data_cache["security"]["data"])
    }


@app.get("/api/remittances")
async def get_remittances():
    """Get workers' remittances data from State Bank of Pakistan"""
    await refresh_cache_with_persistence("remittances", 3600, fetch_remittances_data)
    
    return {
        "data": data_cache["remittances"]["data"],
        "updated": data_cache["remittances"]["updated"].isoformat() if data_cache["remittances"]["updated"] else None
    }

@app.get("/api/gold-reserves")
async def get_gold_reserves():
    """Get gold reserves data from State Bank of Pakistan"""
    await refresh_cache_with_persistence("gold_reserves", 3600, fetch_gold_reserves_data)
    
    return {
        "data": data_cache["gold_reserves"]["data"],
        "updated": data_cache["gold_reserves"]["updated"].isoformat() if data_cache["gold_reserves"]["updated"] else None
    }

@app.get("/api/forex-reserves")
async def get_forex_reserves():
    """Get total forex reserves data from State Bank of Pakistan"""
    await refresh_cache_with_persistence("forex_reserves", 3600, fetch_forex_reserves_data)
    
    return {
        "data": data_cache["forex_reserves"]["data"],
        "updated": data_cache["forex_reserves"]["updated"].isoformat() if data_cache["forex_reserves"]["updated"] else None
    }


@app.get("/api/liquid-forex")
async def get_liquid_forex():
    """Get Liquid Foreign Exchange Reserves data (weekly)"""
    # Refresh every 6 hours (weekly data)
    if data_cache["liquid_forex"]["updated"]:
        age = (datetime.now(timezone.utc) - data_cache["liquid_forex"]["updated"]).total_seconds()
        if age > 21600:  # 6 hours
            data_cache["liquid_forex"]["data"] = await fetch_liquid_forex_data()
            data_cache["liquid_forex"]["updated"] = datetime.now(timezone.utc)
    else:
        data_cache["liquid_forex"]["data"] = await fetch_liquid_forex_data()
        data_cache["liquid_forex"]["updated"] = datetime.now(timezone.utc)
    
    return {
        "data": data_cache["liquid_forex"]["data"],
        "updated": data_cache["liquid_forex"]["updated"].isoformat() if data_cache["liquid_forex"]["updated"] else None
    }


@app.get("/api/current-account")
async def get_current_account():
    """Get current account balance data from State Bank of Pakistan"""
    await refresh_cache_with_persistence("current_account", 3600, fetch_current_account_data)
    
    return {
        "data": data_cache["current_account"]["data"],
        "updated": data_cache["current_account"]["updated"].isoformat() if data_cache["current_account"]["updated"] else None
    }

@app.get("/api/imports")
async def get_imports():
    """Get imports data from State Bank of Pakistan"""
    await refresh_cache_with_persistence("imports", 3600, fetch_imports_data)
    
    return {
        "data": data_cache["imports"]["data"],
        "updated": data_cache["imports"]["updated"].isoformat() if data_cache["imports"]["updated"] else None
    }

@app.get("/api/exports")
async def get_exports():
    """Get exports data from State Bank of Pakistan"""
    await refresh_cache_with_persistence("exports", 3600, fetch_exports_data)
    
    return {
        "data": data_cache["exports"]["data"],
        "updated": data_cache["exports"]["updated"].isoformat() if data_cache["exports"]["updated"] else None
    }


@app.get("/api/fdi")
async def get_fdi():
    """Get Foreign Direct Investment data"""
    await refresh_cache_with_persistence("fdi", 21600, fetch_fdi_data)

    return {
        "data": data_cache["fdi"]["data"],
        "updated": data_cache["fdi"]["updated"].isoformat() if data_cache["fdi"]["updated"] else None
    }


@app.get("/api/gov-debt")
async def get_gov_debt():
    """Get Central Government Debt data (total + internal + external)"""
    await refresh_cache_with_persistence("gov_debt", 21600, fetch_gov_debt_data)

    return {
        "data": data_cache["gov_debt"]["data"],
        "updated": data_cache["gov_debt"]["updated"].isoformat() if data_cache["gov_debt"]["updated"] else None
    }


@app.get("/api/business-environment")
async def get_business_environment():
    """Get EPU + Business Confidence composite dataset for Business Environment panel"""
    # Refresh every 6 hours (monthly survey/index data)
    should_refresh = False
    cached_data = data_cache["business_environment"]["data"]
    if not data_cache["business_environment"]["updated"] or not cached_data:
        should_refresh = True
    else:
        age = (datetime.now(timezone.utc) - data_cache["business_environment"]["updated"]).total_seconds()
        history = cached_data.get("confidence", {}).get("history", []) if isinstance(cached_data, dict) else []
        earliest_date = history[0].get("date") if history else None
        latest_date = history[-1].get("date") if history else None

        now_utc = datetime.now(timezone.utc)
        prev_month_year = now_utc.year if now_utc.month > 1 else now_utc.year - 1
        prev_month = now_utc.month - 1 if now_utc.month > 1 else 12
        expected_latest_min = f"{prev_month_year}-{prev_month:02d}-01"

        should_refresh = (
            age > 21600
            or (earliest_date is not None and earliest_date != "2017-10-01")
            or (latest_date is not None and latest_date < expected_latest_min)
        )

    if should_refresh:
        fetched = await fetch_business_environment_data()
        if fetched:
            data_cache["business_environment"]["data"] = fetched
            data_cache["business_environment"]["updated"] = datetime.now(timezone.utc)
            persist_cache_entry("business_environment", fetched)
        elif not data_cache["business_environment"]["data"]:
            persisted_data, persisted_updated = restore_cache_entry("business_environment")
            if persisted_data:
                data_cache["business_environment"]["data"] = persisted_data
                data_cache["business_environment"]["updated"] = datetime.fromisoformat(persisted_updated) if persisted_updated else datetime.now(timezone.utc)
    elif not data_cache["business_environment"]["data"]:
        persisted_data, persisted_updated = restore_cache_entry("business_environment")
        if persisted_data:
            data_cache["business_environment"]["data"] = persisted_data
            data_cache["business_environment"]["updated"] = datetime.fromisoformat(persisted_updated) if persisted_updated else datetime.now(timezone.utc)

    return {
        "data": data_cache["business_environment"]["data"],
        "updated": data_cache["business_environment"]["updated"].isoformat() if data_cache["business_environment"]["updated"] else None
    }


@app.get("/api/road-advisory")
async def get_road_advisory():
    """Get road advisory data from NHMP"""
    # Refresh every 30 minutes
    if data_cache["road_advisory"]["updated"]:
        age = (datetime.now(timezone.utc) - data_cache["road_advisory"]["updated"]).total_seconds()
        if age > 1800:  # 30 minutes
            data_cache["road_advisory"]["data"] = await fetch_road_advisory()
            data_cache["road_advisory"]["updated"] = datetime.now(timezone.utc)
    else:
        data_cache["road_advisory"]["data"] = await fetch_road_advisory()
        data_cache["road_advisory"]["updated"] = datetime.now(timezone.utc)
    
    return {
        "advisories": data_cache["road_advisory"]["data"],
        "count": len(data_cache["road_advisory"]["data"]),
        "source": "National Highway & Motorway Police",
        "updated": data_cache["road_advisory"]["updated"].isoformat() if data_cache["road_advisory"]["updated"] else None
    }


@app.get("/api/pkr-usd")
async def get_pkr_usd():
    """Get PKR/USD exchange rate from State Bank of Pakistan"""
    await refresh_cache_with_persistence("pkr_usd", 3600, fetch_pkr_usd_data)
    
    return {
        "data": data_cache["pkr_usd"]["data"],
        "updated": data_cache["pkr_usd"]["updated"].isoformat() if data_cache["pkr_usd"]["updated"] else None
    }


@app.get("/api/cpi-yoy")
async def get_cpi_yoy():
    """Get CPI Year-on-Year inflation data from State Bank of Pakistan"""
    # Refresh every hour (monthly data)
    if data_cache["cpi_yoy"]["updated"]:
        age = (datetime.now(timezone.utc) - data_cache["cpi_yoy"]["updated"]).total_seconds()
        if age > 3600:
            data_cache["cpi_yoy"]["data"] = await fetch_cpi_data("yoy")
            data_cache["cpi_yoy"]["updated"] = datetime.now(timezone.utc)
    else:
        data_cache["cpi_yoy"]["data"] = await fetch_cpi_data("yoy")
        data_cache["cpi_yoy"]["updated"] = datetime.now(timezone.utc)
    
    return {
        "data": data_cache["cpi_yoy"]["data"],
        "updated": data_cache["cpi_yoy"]["updated"].isoformat() if data_cache["cpi_yoy"]["updated"] else None
    }


@app.get("/api/cpi-mom")
async def get_cpi_mom():
    """Get CPI Month-on-Month inflation data from State Bank of Pakistan"""
    # Refresh every hour (monthly data)
    if data_cache["cpi_mom"]["updated"]:
        age = (datetime.now(timezone.utc) - data_cache["cpi_mom"]["updated"]).total_seconds()
        if age > 3600:
            data_cache["cpi_mom"]["data"] = await fetch_cpi_data("mom")
            data_cache["cpi_mom"]["updated"] = datetime.now(timezone.utc)
    else:
        data_cache["cpi_mom"]["data"] = await fetch_cpi_data("mom")
        data_cache["cpi_mom"]["updated"] = datetime.now(timezone.utc)
    
    return {
        "data": data_cache["cpi_mom"]["data"],
        "updated": data_cache["cpi_mom"]["updated"].isoformat() if data_cache["cpi_mom"]["updated"] else None
    }


@app.get("/api/cpi-yoy-historical")
async def get_cpi_yoy_historical():
    """Get complete historical CPI Year-on-Year inflation data (1964-present)"""
    # Refresh every 24 hours (historical data rarely changes)
    if data_cache["cpi_yoy_historical"]["updated"]:
        age = (datetime.now(timezone.utc) - data_cache["cpi_yoy_historical"]["updated"]).total_seconds()
        if age > 86400:  # 24 hours
            data_cache["cpi_yoy_historical"]["data"] = await fetch_cpi_historical_data("yoy")
            data_cache["cpi_yoy_historical"]["updated"] = datetime.now(timezone.utc)
    else:
        data_cache["cpi_yoy_historical"]["data"] = await fetch_cpi_historical_data("yoy")
        data_cache["cpi_yoy_historical"]["updated"] = datetime.now(timezone.utc)
    
    return {
        "data": data_cache["cpi_yoy_historical"]["data"],
        "updated": data_cache["cpi_yoy_historical"]["updated"].isoformat() if data_cache["cpi_yoy_historical"]["updated"] else None
    }


@app.get("/api/cpi-mom-historical")
async def get_cpi_mom_historical():
    """Get complete historical CPI Month-on-Month inflation data (1964-present)"""
    # Refresh every 24 hours (historical data rarely changes)
    if data_cache["cpi_mom_historical"]["updated"]:
        age = (datetime.now(timezone.utc) - data_cache["cpi_mom_historical"]["updated"]).total_seconds()
        if age > 86400:  # 24 hours
            data_cache["cpi_mom_historical"]["data"] = await fetch_cpi_historical_data("mom")
            data_cache["cpi_mom_historical"]["updated"] = datetime.now(timezone.utc)
    else:
        data_cache["cpi_mom_historical"]["data"] = await fetch_cpi_historical_data("mom")
        data_cache["cpi_mom_historical"]["updated"] = datetime.now(timezone.utc)
    
    return {
        "data": data_cache["cpi_mom_historical"]["data"],
        "updated": data_cache["cpi_mom_historical"]["updated"].isoformat() if data_cache["cpi_mom_historical"]["updated"] else None
    }


async def ensure_lsm_historical_cache():
    """Refresh LSM historical cache safely (without clobbering last good data on API limits)."""
    should_refresh = False
    if not data_cache["lsm_historical"]["updated"] or not data_cache["lsm_historical"]["data"]:
        should_refresh = True
    else:
        age = (datetime.now(timezone.utc) - data_cache["lsm_historical"]["updated"]).total_seconds()
        should_refresh = age > 21600

    if should_refresh:
        fetched = await fetch_lsm_historical_data()
        if fetched:
            data_cache["lsm_historical"]["data"] = fetched
            data_cache["lsm_historical"]["updated"] = datetime.now(timezone.utc)
            persist_cache_entry("lsm_historical", fetched)
        elif not data_cache["lsm_historical"]["data"]:
            persisted_data, persisted_updated = restore_cache_entry("lsm_historical")
            if persisted_data:
                data_cache["lsm_historical"]["data"] = persisted_data
                data_cache["lsm_historical"]["updated"] = datetime.fromisoformat(persisted_updated) if persisted_updated else datetime.now(timezone.utc)
    elif not data_cache["lsm_historical"]["data"]:
        persisted_data, persisted_updated = restore_cache_entry("lsm_historical")
        if persisted_data:
            data_cache["lsm_historical"]["data"] = persisted_data
            data_cache["lsm_historical"]["updated"] = datetime.fromisoformat(persisted_updated) if persisted_updated else datetime.now(timezone.utc)


@app.get("/api/lsm")
async def get_lsm():
    """Get latest LSM quantum index summary data"""
    await ensure_lsm_historical_cache()

    historical = data_cache["lsm_historical"]["data"]
    summary = build_lsm_summary_from_historical(historical) if historical else None

    if summary:
        data_cache["lsm"]["data"] = summary
        data_cache["lsm"]["updated"] = datetime.now(timezone.utc)
        persist_cache_entry("lsm", summary)
    elif not data_cache["lsm"]["data"]:
        persisted_data, persisted_updated = restore_cache_entry("lsm")
        if persisted_data:
            data_cache["lsm"]["data"] = persisted_data
            data_cache["lsm"]["updated"] = datetime.fromisoformat(persisted_updated) if persisted_updated else datetime.now(timezone.utc)

    return {
        "data": data_cache["lsm"]["data"],
        "updated": data_cache["lsm"]["updated"].isoformat() if data_cache["lsm"]["updated"] else None
    }


@app.get("/api/lsm-historical")
async def get_lsm_historical():
    """Get complete historical LSM quantum index data across base periods"""
    await ensure_lsm_historical_cache()

    return {
        "data": data_cache["lsm_historical"]["data"],
        "updated": data_cache["lsm_historical"]["updated"].isoformat() if data_cache["lsm_historical"]["updated"] else None
    }


@app.get("/api/auto-vehicles")
async def get_auto_vehicles():
    """Get production and sales of auto vehicles with category breakdowns"""
    await refresh_cache_with_persistence("auto_vehicles", 21600, fetch_auto_vehicle_data)

    return {
        "data": data_cache["auto_vehicles"]["data"],
        "updated": data_cache["auto_vehicles"]["updated"].isoformat() if data_cache["auto_vehicles"]["updated"] else None
    }


@app.get("/api/fertilizer")
async def get_fertilizer():
    """Get fertilizer sales/offtake data"""
    await refresh_cache_with_persistence("fertilizer", 21600, fetch_fertilizer_data)

    # Ensure full expected historical window exists (Jul 2006 onward)
    cached = data_cache["fertilizer"]["data"] if isinstance(data_cache["fertilizer"]["data"], dict) else None
    history = cached.get("history", []) if cached else []
    earliest_date = history[0].get("date") if history else None
    if earliest_date and earliest_date > "2006-07-01":
        fetched = await fetch_fertilizer_data()
        if fetched:
            data_cache["fertilizer"]["data"] = fetched
            data_cache["fertilizer"]["updated"] = datetime.now(timezone.utc)
            persist_cache_entry("fertilizer", fetched)

    return {
        "data": data_cache["fertilizer"]["data"],
        "updated": data_cache["fertilizer"]["updated"].isoformat() if data_cache["fertilizer"]["updated"] else None
    }


@app.get("/api/spi-weekly")
async def get_spi_weekly():
    """Get weekly SPI data from Google Sheet (main raw data)"""
    # Refresh every 12 hours (weekly source; user requested low-frequency refresh)
    should_refresh = False
    if not data_cache["spi_weekly"]["updated"] or not data_cache["spi_weekly"]["data"]:
        should_refresh = True
    else:
        age = (datetime.now(timezone.utc) - data_cache["spi_weekly"]["updated"]).total_seconds()
        should_refresh = age > 43200

    if should_refresh:
        data_cache["spi_weekly"]["data"] = await fetch_spi_weekly_data()
        data_cache["spi_weekly"]["updated"] = datetime.now(timezone.utc)

    return {
        "data": data_cache["spi_weekly"]["data"],
        "updated": data_cache["spi_weekly"]["updated"].isoformat() if data_cache["spi_weekly"]["updated"] else None
    }


@app.get("/api/spi-monthly")
async def get_spi_monthly():
    """Get monthly SPI data from Google Sheet (monthly_spi)"""
    # Refresh every 12 hours (monthly source; user requested low-frequency refresh)
    should_refresh = False
    if not data_cache["spi_monthly"]["updated"] or not data_cache["spi_monthly"]["data"]:
        should_refresh = True
    else:
        age = (datetime.now(timezone.utc) - data_cache["spi_monthly"]["updated"]).total_seconds()
        should_refresh = age > 43200

    if should_refresh:
        data_cache["spi_monthly"]["data"] = await fetch_spi_monthly_data()
        data_cache["spi_monthly"]["updated"] = datetime.now(timezone.utc)

    return {
        "data": data_cache["spi_monthly"]["data"],
        "updated": data_cache["spi_monthly"]["updated"].isoformat() if data_cache["spi_monthly"]["updated"] else None
    }


async def fetch_psx_data():
    """Fetch PSX KSE-100 data by scraping the PSX data portal"""
    import re
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            response = await client.get(
                "https://dps.psx.com.pk/",
                headers={
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
                }
            )
            
            if response.status_code == 200:
                html = response.text
                
                # Extract KSE100 data from the indices slider
                kse100_match = re.search(
                    r'topIndices__item__name">KSE100</div><div class="topIndices__item__val">([0-9,\.]+)</div></div><div class="change__text--(neg|pos)"><div class="topIndices__item__change"><i class="icon-(down|up)-dir"></i>\s*([-0-9,\.]+)</div><div class="topIndices__item__changep">\(([-0-9\.]+)%\)</div>',
                    html
                )
                
                if kse100_match:
                    value_str = kse100_match.group(1).replace(',', '')
                    change_str = kse100_match.group(4).replace(',', '')
                    change_percent_str = kse100_match.group(5)
                    
                    value = float(value_str)
                    change = float(change_str)
                    change_percent = float(change_percent_str)
                    
                    # Extract detailed stats from the KSE100 panel
                    # Find the KSE100 panel section
                    panel_match = re.search(
                        r'tabs__panel marketIndices__details" data-name="KSE100"[^>]*data-date="([^"]+)"[^>]*>(.*?)</div></div></div></div>',
                        html,
                        re.DOTALL
                    )
                    
                    timestamp = None
                    high = None
                    low = None
                    volume = None
                    yoy_change = None
                    ytd_change = None
                    prev_close = None
                    
                    if panel_match:
                        timestamp = panel_match.group(1)
                        panel_html = panel_match.group(2)
                        
                        # Extract High
                        high_match = re.search(r'stats_label">High</div><div class="stats_value">([0-9,\.]+)</div>', panel_html)
                        if high_match:
                            high = float(high_match.group(1).replace(',', ''))
                        
                        # Extract Low
                        low_match = re.search(r'stats_label">Low</div><div class="stats_value">([0-9,\.]+)</div>', panel_html)
                        if low_match:
                            low = float(low_match.group(1).replace(',', ''))
                        
                        # Extract Volume
                        vol_match = re.search(r'stats_label">Volume</div><div class="stats_value">([0-9,\.]+)</div>', panel_html)
                        if vol_match:
                            volume = int(vol_match.group(1).replace(',', ''))
                        
                        # Extract 1-Year Change
                        yoy_match = re.search(r'stats_label">1-Year Change</div><div class="stats_value change__text--(pos|neg)">[^0-9]*([\d\.]+)%</div>', panel_html)
                        if yoy_match:
                            yoy_change = float(yoy_match.group(2))
                            if yoy_match.group(1) == 'neg':
                                yoy_change = -yoy_change
                        
                        # Extract YTD Change
                        ytd_match = re.search(r'stats_label">YTD Change</div><div class="stats_value change__text--(pos|neg)">.*?([-]?[\d\.]+)%</div>', panel_html)
                        if ytd_match:
                            ytd_change = float(ytd_match.group(2))
                            if ytd_match.group(1) == 'neg' and ytd_change > 0:
                                ytd_change = -ytd_change
                        
                        # Extract Previous Close
                        prev_match = re.search(r'stats_label">Previous Close</div><div class="stats_value">([0-9,\.]+)</div>', panel_html)
                        if prev_match:
                            prev_close = float(prev_match.group(1).replace(',', ''))
                    
                    return {
                        "value": value,
                        "change": change,
                        "change_percent": change_percent,
                        "high": high,
                        "low": low,
                        "volume": volume,
                        "yoy_change": yoy_change,
                        "ytd_change": ytd_change,
                        "previous_close": prev_close,
                        "timestamp": timestamp,
                        "source": "Pakistan Stock Exchange",
                        "updated": datetime.now(timezone.utc).isoformat()
                    }
    except Exception as e:
        print(f"Error fetching PSX data: {e}")
    
    return None


@app.get("/api/psx-data")
async def get_psx_data():
    """Get PSX KSE-100 index data"""
    # Refresh every 5 minutes during market hours
    if data_cache["psx_data"]["updated"]:
        age = (datetime.now(timezone.utc) - data_cache["psx_data"]["updated"]).total_seconds()
        if age > 300:  # 5 minutes
            data_cache["psx_data"]["data"] = await fetch_psx_data()
            data_cache["psx_data"]["updated"] = datetime.now(timezone.utc)
    else:
        data_cache["psx_data"]["data"] = await fetch_psx_data()
        data_cache["psx_data"]["updated"] = datetime.now(timezone.utc)
    
    return {
        "data": data_cache["psx_data"]["data"],
        "updated": data_cache["psx_data"]["updated"].isoformat() if data_cache["psx_data"]["updated"] else None
    }


@app.get("/api/regional")
async def get_regional_relations():
    """Get regional relations data"""
    relations = {
        "india": {
            "status": "Tense",
            "trade_status": "Limited",
            "recent_events": [
                "Border security talks scheduled",
                "Trade discussions on hold"
            ],
            "sentiment": -0.3
        },
        "china": {
            "status": "Strategic Partnership",
            "trade_status": "Active",
            "recent_events": [
                "CPEC Phase 2 projects advancing",
                "Military cooperation exercises planned"
            ],
            "sentiment": 0.8
        },
        "afghanistan": {
            "status": "Complex",
            "trade_status": "Limited",
            "recent_events": [
                "Border management discussions ongoing",
                "Humanitarian aid coordination"
            ],
            "sentiment": 0.1
        },
        "iran": {
            "status": "Cordial",
            "trade_status": "Growing",
            "recent_events": [
                "Energy cooperation talks",
                "Border security coordination"
            ],
            "sentiment": 0.4
        },
        "saudi_arabia": {
            "status": "Strong Alliance",
            "trade_status": "Active",
            "recent_events": [
                "Investment discussions ongoing",
                "Worker welfare agreements"
            ],
            "sentiment": 0.7
        },
        "usa": {
            "status": "Engagement",
            "trade_status": "Normal",
            "recent_events": [
                "Diplomatic engagements continue",
                "Trade discussions scheduled"
            ],
            "sentiment": 0.3
        }
    }
    return {
        "relations": relations,
        "updated": datetime.now(timezone.utc).isoformat()
    }

@app.get("/api/infrastructure")
async def get_infrastructure_status():
    """Get infrastructure monitoring data"""
    # Check if flight data needs refresh (every 6 hours)
    needs_airport_refresh = False
    for airport_key in AIRPORTS.keys():
        if flight_cache[airport_key]["updated"]:
            age = (datetime.now(timezone.utc) - flight_cache[airport_key]["updated"]).total_seconds()
            if age > 21600:  # 6 hours
                needs_airport_refresh = True
                break
        else:
            needs_airport_refresh = True
            break
    
    if needs_airport_refresh:
        await fetch_all_airports()
    
    # Check if port data needs refresh (every 6 hours)
    needs_port_refresh = False
    for port_key in PORTS.keys():
        if port_cache[port_key]["updated"]:
            age = (datetime.now(timezone.utc) - port_cache[port_key]["updated"]).total_seconds()
            if age > 21600:  # 6 hours
                needs_port_refresh = True
                break
        else:
            needs_port_refresh = True
            break
    
    if needs_port_refresh:
        await fetch_all_ports()
    
    # Build airports data
    airports_data = {}
    for key, info in AIRPORTS.items():
        airports_data[key] = {
            "code": info["code"],
            "name": info["name"],
            "departures": flight_cache[key].get("departures", 0),
            "arrivals": flight_cache[key].get("arrivals", 0),
            "departures_url": f"{FLIGHTSTATS_BASE}/departures/{info['code']}",
            "arrivals_url": f"{FLIGHTSTATS_BASE}/arrivals/{info['code']}"
        }
    
    # Build ports data
    ports_data = {}
    for key, info in PORTS.items():
        ports_data[key] = {
            "code": info["code"],
            "name": info["name"],
            "in_port": port_cache[key].get("in_port", 0),
            "arrivals": port_cache[key].get("arrivals", 0),
            "departures": port_cache[key].get("departures", 0),
            "expected": port_cache[key].get("expected", 0),
            "url": f"https://www.myshiptracking.com/ports/port-of-{info['name'].lower().replace(' ', '-')}-in-pk-pakistan-id-{info['id']}"
        }
    
    infrastructure = {
        "power": {
            "national_grid_status": "Stable",
            "load_shedding": {
                "urban": "2-3 hours",
                "rural": "4-6 hours"
            },
            "generation_capacity": "85%",
            "demand": "22,500 MW",
            "supply": "21,000 MW"
        },
        "internet": {
            "national_status": "Operational",
            "average_speed": "12.5 Mbps",
            "outage_reports": 3,
            "affected_regions": ["Parts of Balochistan"]
        },
        "airport_status": {
            "airports": airports_data,
            "note": "24 hours"
        },
        "port_status": {
            "ports": ports_data,
            "note": "24 hours"
        },
        "updated": datetime.now(timezone.utc).isoformat()
    }
    return infrastructure

@app.get("/api/map-data")
async def get_map_data():
    """Get map markers and data points"""
    markers = {
        "cities": [
            {"name": "Islamabad", "lat": 33.6844, "lon": 73.0479, "type": "capital", "population": "1.1M"},
            {"name": "Karachi", "lat": 24.8607, "lon": 67.0011, "type": "major", "population": "14.9M"},
            {"name": "Lahore", "lat": 31.5204, "lon": 74.3587, "type": "major", "population": "11.1M"},
            {"name": "Faisalabad", "lat": 31.4504, "lon": 73.1350, "type": "major", "population": "3.2M"},
            {"name": "Rawalpindi", "lat": 33.5651, "lon": 73.0169, "type": "major", "population": "2.1M"},
            {"name": "Peshawar", "lat": 34.0151, "lon": 71.5249, "type": "major", "population": "1.9M"},
            {"name": "Quetta", "lat": 30.1798, "lon": 66.9750, "type": "provincial", "population": "1.0M"},
            {"name": "Multan", "lat": 30.1575, "lon": 71.5249, "type": "major", "population": "1.8M"},
            {"name": "Hyderabad", "lat": 25.3960, "lon": 68.3578, "type": "major", "population": "1.7M"},
            {"name": "Gwadar", "lat": 25.1264, "lon": 62.3225, "type": "strategic", "population": "0.1M"}
        ],
        "cpec_routes": [
            {"start": [35.9220, 74.3081], "end": [25.1264, 62.3225], "name": "Western Route"},
            {"start": [33.6844, 73.0479], "end": [24.8607, 67.0011], "name": "Central Route"},
            {"start": [31.5204, 74.3587], "end": [24.8607, 67.0011], "name": "Eastern Route"}
        ],
        "center": [30.3753, 69.3451],
        "zoom": 5,
        "updated": datetime.now(timezone.utc).isoformat()
    }
    return markers


# Admin endpoints for historical data
@app.get("/api/admin/airport-history")
async def get_airport_history(
    airport_code: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    limit: int = 100
):
    """Get historical airport data for admin analysis
    
    Args:
        airport_code: Filter by airport code (ISB, KHI, LHE, PEW, MUX)
        start_date: Start date filter (YYYY-MM-DD)
        end_date: End date filter (YYYY-MM-DD)
        limit: Maximum records to return (default 100, max 1000)
    """
    query = {}
    
    if airport_code:
        query["airport_code"] = airport_code.upper()
    
    if start_date:
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            query["timestamp"] = {"$gte": start_dt.replace(tzinfo=timezone.utc)}
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid start_date format. Use YYYY-MM-DD")
    
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
            if "timestamp" in query:
                query["timestamp"]["$lt"] = end_dt.replace(tzinfo=timezone.utc)
            else:
                query["timestamp"] = {"$lt": end_dt.replace(tzinfo=timezone.utc)}
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid end_date format. Use YYYY-MM-DD")
    
    limit = min(limit, 1000)
    
    records = list(airport_history_collection.find(
        query, 
        {"_id": 0}
    ).sort("timestamp", -1).limit(limit))
    
    # Convert datetime objects to ISO strings
    for record in records:
        if isinstance(record.get("timestamp"), datetime):
            record["timestamp"] = record["timestamp"].isoformat()
    
    return {
        "total_records": airport_history_collection.count_documents(query),
        "returned_records": len(records),
        "data": records,
        "filters_applied": {
            "airport_code": airport_code,
            "start_date": start_date,
            "end_date": end_date
        }
    }


@app.get("/api/admin/port-history")
async def get_port_history(
    port_code: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    limit: int = 100
):
    """Get historical port data for admin analysis
    
    Args:
        port_code: Filter by port code (PKKHI, PKBQM, PKGWD)
        start_date: Start date filter (YYYY-MM-DD)
        end_date: End date filter (YYYY-MM-DD)
        limit: Maximum records to return (default 100, max 1000)
    """
    query = {}
    
    if port_code:
        query["port_code"] = port_code.upper()
    
    if start_date:
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            query["timestamp"] = {"$gte": start_dt.replace(tzinfo=timezone.utc)}
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid start_date format. Use YYYY-MM-DD")
    
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
            if "timestamp" in query:
                query["timestamp"]["$lt"] = end_dt.replace(tzinfo=timezone.utc)
            else:
                query["timestamp"] = {"$lt": end_dt.replace(tzinfo=timezone.utc)}
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid end_date format. Use YYYY-MM-DD")
    
    limit = min(limit, 1000)
    
    records = list(port_history_collection.find(
        query, 
        {"_id": 0}
    ).sort("timestamp", -1).limit(limit))
    
    # Convert datetime objects to ISO strings
    for record in records:
        if isinstance(record.get("timestamp"), datetime):
            record["timestamp"] = record["timestamp"].isoformat()
    
    return {
        "total_records": port_history_collection.count_documents(query),
        "returned_records": len(records),
        "data": records,
        "filters_applied": {
            "port_code": port_code,
            "start_date": start_date,
            "end_date": end_date
        }
    }


@app.get("/api/admin/airport-history/summary")
async def get_airport_history_summary():
    """Get summary statistics of airport historical data"""
    pipeline = [
        {
            "$group": {
                "_id": "$airport_code",
                "airport_name": {"$first": "$airport_name"},
                "total_records": {"$sum": 1},
                "first_record": {"$min": "$timestamp"},
                "last_record": {"$max": "$timestamp"},
                "avg_departures": {"$avg": "$departures"},
                "avg_arrivals": {"$avg": "$arrivals"},
                "max_departures": {"$max": "$departures"},
                "max_arrivals": {"$max": "$arrivals"}
            }
        },
        {"$sort": {"_id": 1}}
    ]
    
    results = list(airport_history_collection.aggregate(pipeline))
    
    for result in results:
        if isinstance(result.get("first_record"), datetime):
            result["first_record"] = result["first_record"].isoformat()
        if isinstance(result.get("last_record"), datetime):
            result["last_record"] = result["last_record"].isoformat()
        if result.get("avg_departures"):
            result["avg_departures"] = round(result["avg_departures"], 1)
        if result.get("avg_arrivals"):
            result["avg_arrivals"] = round(result["avg_arrivals"], 1)
    
    return {
        "total_airports": len(results),
        "total_records": airport_history_collection.count_documents({}),
        "airports": results
    }


@app.get("/api/admin/port-history/summary")
async def get_port_history_summary():
    """Get summary statistics of port historical data"""
    pipeline = [
        {
            "$group": {
                "_id": "$port_code",
                "port_name": {"$first": "$port_name"},
                "total_records": {"$sum": 1},
                "first_record": {"$min": "$timestamp"},
                "last_record": {"$max": "$timestamp"},
                "avg_in_port": {"$avg": "$in_port"},
                "avg_arrivals": {"$avg": "$arrivals"},
                "avg_departures": {"$avg": "$departures"},
                "max_in_port": {"$max": "$in_port"},
                "max_arrivals": {"$max": "$arrivals"}
            }
        },
        {"$sort": {"_id": 1}}
    ]
    
    results = list(port_history_collection.aggregate(pipeline))
    
    for result in results:
        if isinstance(result.get("first_record"), datetime):
            result["first_record"] = result["first_record"].isoformat()
        if isinstance(result.get("last_record"), datetime):
            result["last_record"] = result["last_record"].isoformat()
        if result.get("avg_in_port"):
            result["avg_in_port"] = round(result["avg_in_port"], 1)
        if result.get("avg_arrivals"):
            result["avg_arrivals"] = round(result["avg_arrivals"], 1)
        if result.get("avg_departures"):
            result["avg_departures"] = round(result["avg_departures"], 1)
    
    return {
        "total_ports": len(results),
        "total_records": port_history_collection.count_documents({}),
        "ports": results
    }


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)
